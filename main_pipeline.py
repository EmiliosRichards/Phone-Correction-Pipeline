import pandas as pd
from typing import List, Dict, Set, Optional, Any, Callable, Union, Tuple
from collections import Counter # Added for duplicate counting
import csv # Added for failure log
from src.data_handler import load_and_preprocess_data, process_and_consolidate_contact_data, get_canonical_base_url, generate_processed_contacts_report # Kept main's import
from src.scraper import scrape_website
from src.regex_extractor_component import extract_numbers_with_snippets_from_text
from src.llm_extractor_component import GeminiLLMExtractor
from src.core.schemas import PhoneNumberLLMOutput, CompanyContactDetails, ConsolidatedPhoneNumber 
from src.scraper.scraper_logic import normalize_url
from src.core.logging_config import setup_logging
from src.core.config import AppConfig
import logging
import os
import asyncio
from datetime import datetime
import time
import json
import re
from urllib.parse import urlparse, quote
import socket # Added for TLD probing
import phonenumbers
from phonenumbers import NumberParseException
from openpyxl.utils import get_column_letter

TARGET_COUNTRY_CODES_INT: Set[int] = {49, 41, 43} 
EXCLUDED_TYPES_FOR_TOP_CONTACTS_REPORT: Set[str] = {
    'Unknown', 'Fax', 'Mobile', 'Date', 'ID' 
}
logger = logging.getLogger(__name__) 
app_config: AppConfig = AppConfig()

FAULT_CATEGORY_MAP_DEFINITION: Dict[str, str] = {
    "Input_URL_Invalid": "Input Data Issue",
    "Input_URL_UnsupportedScheme": "Input Data Issue",
    "Scraping_AllAttemptsFailed_Network": "Website Issue",
    "Scraping_AllAttemptsFailed_AccessDenied": "Website Issue",
    "Scraping_ContentNotFound_AllAttempts": "Website Issue",
    "Scraping_Success_NoRelevantContentPagesFound": "Website Issue",
    "Canonical_Duplicate_SkippedProcessing": "Pipeline Logic/Configuration",
    "Canonical_NoRegexCandidatesFound": "Pipeline Logic/Configuration",
    "LLM_NoInput_NoRegexCandidates": "Pipeline Logic/Configuration",
    "LLM_Output_NoNumbersFound_AllAttempts": "LLM Issue",
    "LLM_Output_NumbersFound_NoneRelevant_AllAttempts": "LLM Issue",
    "LLM_Processing_Error_AllAttempts": "LLM Issue",
    "DataConsolidation_Error_ForRow": "Pipeline Error",
    "Pipeline_Skipped_MaxRedirects_ForInputURL": "Website Issue",
    "Pipeline_Skipped_PreviouslyFailedInput": "Pipeline Logic/Configuration", # For future use
    "Unknown_Processing_Gap_NoContact": "Unknown"
}
INPUT_FILE_PATH: str = app_config.input_excel_file_path
if not os.path.isabs(INPUT_FILE_PATH):
    project_root_dir = os.path.dirname(os.path.abspath(__file__))
    INPUT_FILE_PATH = os.path.join(project_root_dir, INPUT_FILE_PATH)
    print(f"INFO: Resolved relative INPUT_FILE_PATH to absolute: {INPUT_FILE_PATH}")


def is_target_country_number_reliable(phone_number_str: str) -> bool:
    if not phone_number_str or not isinstance(phone_number_str, str):
        return False
    try:
        parsed_num = phonenumbers.parse(phone_number_str, None)
        return parsed_num.country_code in TARGET_COUNTRY_CODES_INT
    except NumberParseException:
        logger.debug(f"NumberParseException for '{phone_number_str}' during target country check.")
        return False

def generate_run_id() -> str:
    return datetime.now().strftime("%Y%m%d_%H%M%S")

def log_row_failure(
    failure_log_writer: Optional[Any],
    input_row_identifier: Any,
    company_name: str,
    given_url: Optional[str],
    stage_of_failure: str,
    error_reason: str,
    log_timestamp: str,
    error_details: str = "",
    associated_pathful_canonical_url: Optional[str] = None # New parameter
) -> None:
    """Helper function to write a row-specific failure to the CSV log."""
    if failure_log_writer:
        try:
            sanitized_reason = str(error_reason).replace('\n', ' ').replace('\r', '')
            sanitized_details = str(error_details).replace('\n', ' ').replace('\r', '')
            row_to_write = [
                log_timestamp,
                input_row_identifier,
                company_name,
                given_url if given_url is not None else "",
                stage_of_failure,
                sanitized_reason,
                sanitized_details,
                associated_pathful_canonical_url if associated_pathful_canonical_url is not None else "" # New field
            ]
            failure_log_writer.writerow(row_to_write)
        except Exception as e:
            logger.error(f"CRITICAL: Failed to write to failure_log_csv: {e}. Row ID: {input_row_identifier}, Stage: {stage_of_failure}, Timestamp: {log_timestamp}", exc_info=True)
    else:
        logger.warning(f"Attempted to log row failure but failure_log_writer is None. Row ID: {input_row_identifier}, Stage: {stage_of_failure}, Timestamp: {log_timestamp}")


def get_input_canonical_url(url_string: Optional[str]) -> Optional[str]:
    """
    Normalizes a given URL string and extracts its netloc (domain part)
    to serve as an 'input canonical URL' for pre-computation duplicate checks.
    """
    if not url_string or not isinstance(url_string, str):
        return None
    
    temp_url = url_string.strip()
    if not temp_url: # Handle empty string after strip
        return None

    # Ensure a scheme is present for consistent parsing by normalize_url and for netloc extraction
    parsed_initial = urlparse(temp_url)
    if not parsed_initial.scheme:
        # logger.debug(f"Input URL '{temp_url}' is schemeless for canonical pre-check. Adding 'http://'.") # Optional: add if more detailed logging is needed here
        temp_url = "http://" + temp_url
    
    try:
        # normalize_url is imported from src.scraper.scraper_logic
        normalized = normalize_url(temp_url) # Call without the incorrect parameter
        if not normalized:
            logger.debug(f"normalize_url returned None for input: {temp_url} (original: {url_string})")
            return None
        parsed_normalized_url = urlparse(normalized)
        # Return netloc (domain) if it exists, otherwise None.
        # This ensures that even if normalize_url returns something like "example.com" (no scheme),
        # urlparse().netloc will be empty, which is a valid outcome for a malformed/unresolvable input.
        return parsed_normalized_url.netloc if parsed_normalized_url.netloc else None
    except Exception as e:
        logger.debug(f"Could not parse/normalize input URL '{url_string}' for canonical pre-check: {e}")
        return None

def _determine_final_row_outcome_and_fault(
    index: Any,
    row_summary: pd.Series,
    df_status_snapshot: Dict[str, Any], # Pass relevant df statuses for this row
    company_contact_details_summary: Optional[CompanyContactDetails],
    unique_sorted_consolidated_numbers: List[ConsolidatedPhoneNumber],
    canonical_url_summary: Optional[str], # This is the true_base_domain for the input row
    true_base_scraper_status_map: Dict[str, str], # Map of true_base_domain to its overall scraper status
    true_base_to_pathful_map: Dict[str, List[str]],
    canonical_site_pathful_scraper_status: Dict[str, str], # Map of pathful_url to its scraper status
    canonical_site_raw_llm_outputs: Dict[str, List[PhoneNumberLLMOutput]],
    canonical_site_regex_candidates_found_status: Dict[str, bool], # New
    canonical_site_llm_exception_details: Dict[str, str] # New: For specific LLM error messages
) -> Tuple[str, str]:
    """
    Determines the final outcome reason for an input row and its fault category.
    """
    # Priority 1: Initial Input URL issues (from earlier in the loop)
    # df_status_snapshot should contain 'ScrapingStatus' set during initial URL processing for the row
    initial_row_scrape_status = df_status_snapshot.get('ScrapingStatus', '')
    if initial_row_scrape_status == 'InvalidURL': # This status is set if URL is fundamentally unprocessable
        return "Input_URL_Invalid", FAULT_CATEGORY_MAP_DEFINITION["Input_URL_Invalid"]
    if initial_row_scrape_status == 'MaxRedirects_InputURL': # Assuming this status is set if max redirects hit for input
        return "Pipeline_Skipped_MaxRedirects_ForInputURL", FAULT_CATEGORY_MAP_DEFINITION["Pipeline_Skipped_MaxRedirects_ForInputURL"]
    # Add other direct input URL failure checks here if they are set on df.at[index, 'ScrapingStatus'] early

    # If a contact was successfully extracted for this input row's canonical URL
    if unique_sorted_consolidated_numbers:
        return "Contact_Successfully_Extracted", "N/A"

    # If no canonical URL was determined for the input row (e.g., initial URL was invalid and didn't even lead to a scrape attempt)
    if not canonical_url_summary:
        # This case should ideally be caught by initial_row_scrape_status checks.
        # If it wasn't (e.g. scraper returned no canonical_url but status wasn't 'InvalidURL'),
        # it's a gap or a specific type of scraping failure for the input.
        if initial_row_scrape_status and initial_row_scrape_status != "Success" and initial_row_scrape_status != "Not_Run":
             # Use a more specific scraping failure if available from the initial pass
            return f"ScrapingFailure_InputURL_{initial_row_scrape_status}", FAULT_CATEGORY_MAP_DEFINITION.get(f"Scraping_AllAttemptsFailed_Network", "Website Issue") # Default category
        return "Unknown_NoCanonicalURLDetermined", FAULT_CATEGORY_MAP_DEFINITION["Unknown_Processing_Gap_NoContact"]


    # Check status of the canonical URL
    scraper_status_for_true_base = true_base_scraper_status_map.get(canonical_url_summary, "Unknown")

    if scraper_status_for_true_base != "Success":
        # Determine more specific scraping failure for the canonical URL
        # This requires inspecting canonical_site_pathful_scraper_status for all pathfuls under canonical_url_summary
        pathful_urls_for_canonical = true_base_to_pathful_map.get(canonical_url_summary, [])
        all_network_fail = True
        all_access_denied = True
        all_not_found = True
        if not pathful_urls_for_canonical: # No pathful URLs means scraping didn't really happen for this canonical
             return "Scraping_NoPathfulURLs_ForCanonical", FAULT_CATEGORY_MAP_DEFINITION.get("Scraping_AllAttemptsFailed_Network", "Website Issue")


        for p_url in pathful_urls_for_canonical:
            p_status = canonical_site_pathful_scraper_status.get(p_url, "Unknown")
            if "Timeout" not in p_status and "DNS" not in p_status and "Network" not in p_status and "Unreachable" not in p_status : all_network_fail = False
            if "403" not in p_status and "AccessDenied" not in p_status and "Robots" not in p_status: all_access_denied = False
            if "404" not in p_status and "NotFound" not in p_status : all_not_found = False
        
        if all_network_fail: return "Scraping_AllAttemptsFailed_Network", FAULT_CATEGORY_MAP_DEFINITION["Scraping_AllAttemptsFailed_Network"]
        if all_access_denied: return "Scraping_AllAttemptsFailed_AccessDenied", FAULT_CATEGORY_MAP_DEFINITION["Scraping_AllAttemptsFailed_AccessDenied"]
        if all_not_found: return "Scraping_ContentNotFound_AllAttempts", FAULT_CATEGORY_MAP_DEFINITION["Scraping_ContentNotFound_AllAttempts"]
        # Fallback generic scrape failure for canonical if not fitting above
        return f"ScrapingFailed_Canonical_{scraper_status_for_true_base}", FAULT_CATEGORY_MAP_DEFINITION.get("Scraping_AllAttemptsFailed_Network", "Website Issue")


    # Scraping for canonical was "Success", now check LLM stages
    if initial_row_scrape_status and "Already_Processed" in initial_row_scrape_status: # Check if this input row led to an already processed canonical
        return "Canonical_Duplicate_SkippedProcessing", FAULT_CATEGORY_MAP_DEFINITION["Canonical_Duplicate_SkippedProcessing"]

    # company_contact_details_summary is for the canonical_url_summary

    # Check if regex found candidates for this canonical URL
    if not canonical_site_regex_candidates_found_status.get(canonical_url_summary, False): # Default to False if missing
        return "Canonical_NoRegexCandidatesFound", FAULT_CATEGORY_MAP_DEFINITION["Canonical_NoRegexCandidatesFound"]

    # Regex candidates were found, so LLM should have been attempted or an error occurred
    if company_contact_details_summary is None:
        # This implies an LLM error for the canonical, or prompt missing
        if scraper_status_for_true_base == "Error_LLM_PromptMissing":
            return "LLM_Processing_Error_AllAttempts", FAULT_CATEGORY_MAP_DEFINITION["LLM_Processing_Error_AllAttempts"] # Consider specific "LLM_Prompt_Missing_For_Canonical"
        if scraper_status_for_true_base == "Error_LLM_Processing":
            return "LLM_Processing_Error_AllAttempts", FAULT_CATEGORY_MAP_DEFINITION["LLM_Processing_Error_AllAttempts"]
        # If it reached here, it means regex had candidates, but LLM processing didn't result in a CompanyContactDetails object
        # and it wasn't a logged LLM error for the canonical. This is unusual.
        return "LLM_NoInput_NoRegexCandidates", FAULT_CATEGORY_MAP_DEFINITION["LLM_NoInput_NoRegexCandidates"] # This might be miscategorized now, more like an unknown LLM issue

    # company_contact_details_summary exists, but no consolidated_numbers (LLM ran, results empty or all filtered)
    if not company_contact_details_summary.consolidated_numbers: # Removed redundant company_contact_details_summary check
        # Differentiate based on whether raw LLM output was empty vs. filtered
        all_raw_llm_empty_for_canonical = True
        pathful_urls_for_canonical = true_base_to_pathful_map.get(canonical_url_summary, [])
        if not pathful_urls_for_canonical: # Should not happen if scrape was success
            all_raw_llm_empty_for_canonical = True
        else:
            for p_url in pathful_urls_for_canonical:
                if canonical_site_raw_llm_outputs.get(p_url): # If any pathful URL had non-empty raw LLM output list
                    all_raw_llm_empty_for_canonical = False
                    break
        
        if all_raw_llm_empty_for_canonical:
            return "LLM_Output_NoNumbersFound_AllAttempts", FAULT_CATEGORY_MAP_DEFINITION["LLM_Output_NoNumbersFound_AllAttempts"]
        else:
            return "LLM_Output_NumbersFound_NoneRelevant_AllAttempts", FAULT_CATEGORY_MAP_DEFINITION["LLM_Output_NumbersFound_NoneRelevant_AllAttempts"]

    return "Unknown_Processing_Gap_NoContact", FAULT_CATEGORY_MAP_DEFINITION["Unknown_Processing_Gap_NoContact"]

def _determine_final_domain_outcome_and_fault(
    true_base_domain: str,
    domain_journey_entry: Dict[str, Any], # The entry from canonical_domain_journey_data
    true_base_scraper_status_map: Dict[str, str], # Map of true_base_domain to its overall scraper status
    true_base_to_pathful_map: Dict[str, List[str]], # Maps true_base to list of its pathful URLs
    canonical_site_pathful_scraper_status: Dict[str, str], # Map of pathful_url to its scraper status
    canonical_site_regex_candidates_found_status: Dict[str, bool], # Aggregated: if any pathful under this true_base had regex candidates
    final_consolidated_data: Optional[CompanyContactDetails] # Result of process_and_consolidate_contact_data for this true_base
) -> Tuple[str, str]:
    """
    Determines the final outcome reason for a canonical domain and its primary fault category.
    This is similar to _determine_final_row_outcome_and_fault but operates on domain-level aggregated data.
    """
    # Priority 1: Successful contact extraction for this domain
    if final_consolidated_data and final_consolidated_data.consolidated_numbers:
        return "Contact_Successfully_Extracted_For_Domain", "N/A"

    # Priority 2: Scraper status for the true_base_domain
    # The domain_journey_entry["Overall_Scraper_Status_For_Domain"] should reflect the best status.
    overall_scraper_status_for_domain = domain_journey_entry.get("Overall_Scraper_Status_For_Domain", "Unknown")

    if overall_scraper_status_for_domain != "Success":
        # Try to get more specific scraping failure reasons by looking at pathful statuses
        # This logic mirrors parts of _determine_final_row_outcome_and_fault's scraping failure analysis
        pathful_urls_for_this_domain = true_base_to_pathful_map.get(true_base_domain, [])
        all_network_fail = True
        all_access_denied = True
        all_not_found = True # Or content not found

        if not pathful_urls_for_this_domain:
             # This implies the domain itself might have been invalid or unresolvable before pathfuls were attempted
             # Or, scraper_logic decided not to proceed with any pathfuls.
            if "InvalidURL" in overall_scraper_status_for_domain or "MaxRedirects" in overall_scraper_status_for_domain: # Check for specific input-like errors
                 return f"Domain_InputLikeFailure_{overall_scraper_status_for_domain}", FAULT_CATEGORY_MAP_DEFINITION.get("Input_URL_Invalid", "Input Data Issue")
            return "Scraping_NoPathfulURLs_Processed_ForDomain", FAULT_CATEGORY_MAP_DEFINITION.get("Scraping_AllAttemptsFailed_Network", "Website Issue")


        for p_url in pathful_urls_for_this_domain:
            p_status = canonical_site_pathful_scraper_status.get(p_url, "Unknown")
            if "Timeout" not in p_status and "DNS" not in p_status and "Network" not in p_status and "Unreachable" not in p_status : all_network_fail = False
            if "403" not in p_status and "AccessDenied" not in p_status and "Robots" not in p_status: all_access_denied = False
            if "404" not in p_status and "NotFound" not in p_status : all_not_found = False
        
        if all_network_fail: return "Scraping_AllPathfulsFailed_Network_ForDomain", FAULT_CATEGORY_MAP_DEFINITION["Scraping_AllAttemptsFailed_Network"]
        if all_access_denied: return "Scraping_AllPathfulsFailed_AccessDenied_ForDomain", FAULT_CATEGORY_MAP_DEFINITION["Scraping_AllAttemptsFailed_AccessDenied"]
        if all_not_found: return "Scraping_AllPathfuls_ContentNotFound_ForDomain", FAULT_CATEGORY_MAP_DEFINITION["Scraping_ContentNotFound_AllAttempts"]
        
        # Fallback generic scrape failure for the domain if not fitting above and overall status wasn't success
        return f"ScrapingFailed_Domain_{overall_scraper_status_for_domain}", FAULT_CATEGORY_MAP_DEFINITION.get("Scraping_AllAttemptsFailed_Network", "Website Issue")

    # Scraping for domain was "Success" (at least one pathful URL was successful)
    # Now check Regex and LLM stages based on aggregated data in domain_journey_entry

    if not domain_journey_entry.get("Regex_Candidates_Found_For_Any_Pathful", False):
        # This implies no regex candidates were found on ANY successfully scraped page for this domain.
        # Check if any pages were actually scraped successfully and had content.
        if domain_journey_entry.get("Total_Pages_Scraped_For_Domain", 0) == 0:
             return "Scraping_Success_NoPagesScraped_ForDomain", FAULT_CATEGORY_MAP_DEFINITION["Scraping_Success_NoRelevantContentPagesFound"]
        return "Domain_NoRegexCandidatesFound_OnAnyPage", FAULT_CATEGORY_MAP_DEFINITION["Canonical_NoRegexCandidatesFound"]

    # Regex candidates were found on at least one page for this domain.
    # LLM should have been attempted for those pathful URLs.
    # final_consolidated_data being None or having no consolidated_numbers indicates LLM issues.

    if not domain_journey_entry.get("LLM_Calls_Made_For_Domain", False):
        # This is unusual if regex candidates were found. Could mean an error before LLM call for all relevant pathfuls.
        return "LLM_NotCalled_DespiteRegexCandidates_ForDomain", FAULT_CATEGORY_MAP_DEFINITION["LLM_NoInput_NoRegexCandidates"] # Or a pipeline error

    if domain_journey_entry.get("LLM_Processing_Error_Encountered_For_Domain", False):
        return "LLM_Processing_Error_Encountered_ForDomain", FAULT_CATEGORY_MAP_DEFINITION["LLM_Processing_Error_AllAttempts"]

    # LLM ran, no errors reported at the domain journey level, but still no consolidated numbers.
    # This implies either LLM found nothing, or found nothing relevant after consolidation.
    if domain_journey_entry.get("LLM_Total_Raw_Numbers_Extracted", 0) == 0:
        return "LLM_Output_NoRawNumbersFound_ForDomain", FAULT_CATEGORY_MAP_DEFINITION["LLM_Output_NoNumbersFound_AllAttempts"]
    else: # Raw numbers were extracted, but none made it through consolidation for the domain
        return "LLM_Output_RawNumbersFound_NoneConsolidated_ForDomain", FAULT_CATEGORY_MAP_DEFINITION["LLM_Output_NumbersFound_NoneRelevant_AllAttempts"]

    # Fallback if none of the above conditions met but still no consolidated contacts
    return "Unknown_Domain_Processing_Gap_NoContact", FAULT_CATEGORY_MAP_DEFINITION["Unknown_Processing_Gap_NoContact"]
def main() -> None:
    pipeline_start_time = time.time() 
    run_metrics: Dict[str, Any] = {
        "run_id": None,
        "total_duration_seconds": None,
        "tasks": {},
        "data_processing_stats": {
            "input_rows_count": 0,
            "rows_successfully_processed_pass1": 0,
            "rows_failed_pass1": 0,
            "row_level_failure_summary": {}, # New: For stage_of_failure counts
            "input_unique_company_names": 0,
            "input_unique_canonical_urls": 0,
            "input_company_names_with_duplicates_count": 0, # Number of company names that appear more than once
            "input_canonical_urls_with_duplicates_count": 0, # Number of input canonical URLs that appear more than once
            "input_rows_with_duplicate_company_name": 0, # Total rows that have a company name seen elsewhere
            "input_rows_with_duplicate_canonical_url": 0, # Total rows that have an input canonical URL seen elsewhere
            "input_rows_considered_duplicates_overall": 0, # Total rows where either company name or URL is a duplicate
        },
        "scraping_stats": {
            "urls_processed_for_scraping": 0,
            "scraping_success": 0,
            "scraping_failure_invalid_url": 0,
            "scraping_failure_already_processed": 0,
            "scraping_failure_error": 0, 
            "new_canonical_sites_scraped": 0, 
            "total_pages_scraped_overall": 0,
            "pages_scraped_by_type": {}, 
            "total_successful_canonical_scrapes": 0, 
            "total_urls_fetched_by_scraper": 0, 
        },
        "regex_extraction_stats": {
            "sites_processed_for_regex": 0, 
            "sites_with_regex_candidates": 0,
            "total_regex_candidates_found": 0,
        },
        "llm_processing_stats": {
            "sites_processed_for_llm": 0, 
            "llm_calls_success": 0,
            "llm_calls_failure_prompt_missing": 0,
            "llm_calls_failure_processing_error": 0,
            "llm_no_candidates_to_process": 0, 
            "total_llm_extracted_numbers_raw": 0, 
            "total_llm_prompt_tokens": 0,
            "total_llm_completion_tokens": 0,
            "total_llm_tokens_overall": 0,
            "llm_successful_calls_with_token_data": 0,
        },
        "report_generation_stats": {
            "detailed_report_rows": 0,
            "summary_report_rows": 0,
            "tertiary_report_rows": 0,
            "canonical_domain_summary_rows": 0, # New report
        },
        "errors_encountered": []
    }

    run_id = generate_run_id()
    run_metrics["run_id"] = run_id
    
    output_base_dir_abs: str = app_config.output_base_dir
    if not os.path.isabs(output_base_dir_abs):
        project_root_dir_local = os.path.dirname(os.path.abspath(__file__))
        output_base_dir_abs = os.path.join(project_root_dir_local, output_base_dir_abs)
        
    run_output_dir: str = os.path.join(output_base_dir_abs, run_id)
    os.makedirs(run_output_dir, exist_ok=True)
    
    llm_context_dir = os.path.join(run_output_dir, app_config.llm_context_subdir)
    os.makedirs(llm_context_dir, exist_ok=True)
    
    log_file_name = f"pipeline_run_{run_id}.log"
    log_file_path = os.path.join(run_output_dir, log_file_name)
    
    file_log_level_int = getattr(logging, app_config.log_level.upper(), logging.INFO)
    console_log_level_int = getattr(logging, app_config.console_log_level.upper(), logging.WARNING)
    print(f"DEBUG: main_pipeline.py - Effective console_log_level_int: {console_log_level_int} ({logging.getLevelName(console_log_level_int)})")
    print(f"DEBUG: main_pipeline.py - AppConfig console_log_level raw value: '{app_config.console_log_level}'")
    setup_logging(
        file_log_level=file_log_level_int,
        console_log_level=console_log_level_int,
        log_file_path=log_file_path
    )
    
    logger.info(f"Logging initialized. Run ID: {run_id}")
    logger.info(f"File log level set to: {logging.getLevelName(file_log_level_int)} (from LOG_LEVEL='{app_config.log_level}')")
    logger.info(f"Console log level set to: {logging.getLevelName(console_log_level_int)} (from CONSOLE_LOG_LEVEL='{app_config.console_log_level}')")
    logger.info(f"Main log file will be: {log_file_path}")
    logger.info(f"Base output directory for this run: {run_output_dir}")

    failure_log_csv_path = os.path.join(run_output_dir, f"failed_rows_{run_id}.csv")
    logger.info(f"Row-specific failure log for this run will be: {failure_log_csv_path}")

    logger.info("Starting phone validation pipeline...")
    if not os.path.exists(INPUT_FILE_PATH):
        logger.error(f"CRITICAL: Input file not found at resolved path: {INPUT_FILE_PATH}. Exiting.")
        return

    llm_extractor: GeminiLLMExtractor
    try:
        llm_extractor = GeminiLLMExtractor(config=app_config)
        logger.info("GeminiLLMExtractor initialized successfully.")
    except ValueError as ve:
        logger.error(f"Failed to initialize GeminiLLMExtractor: {ve}. Check GEMINI_API_KEY.")
        return
    except Exception as e:
        logger.error(f"Unexpected error initializing GeminiLLMExtractor: {e}", exc_info=True)
        return

    df: Optional[pd.DataFrame] = None
    task_start_time = time.time()
    try:
        logger.info(f"Attempting to load data from: {INPUT_FILE_PATH}")
        df = load_and_preprocess_data(INPUT_FILE_PATH, app_config_instance=app_config)
        if df is not None:
            logger.info(f"Successfully loaded and preprocessed data from {INPUT_FILE_PATH}. Shape: {df.shape}")
            logger.info(f"DataFrame columns: {df.columns.tolist()}") 
            run_metrics["data_processing_stats"]["input_rows_count"] = len(df)
            if 'GivenURL' in df.columns:
                logger.info(f"First 5 'GivenURL' values: {df['GivenURL'].head().tolist()}")
            else:
                logger.warning("'GivenURL' column not found in the loaded DataFrame.")
                logger.info(f"First 2 rows of loaded DataFrame for inspection:\n{df.head(2).to_string()}")
            logger.debug(f"Loaded DataFrame head:\n{df.head().to_string()}")
        else:
            logger.error(f"Failed to load data from {INPUT_FILE_PATH}. DataFrame is None.")
            run_metrics["errors_encountered"].append(f"Data loading failed: DataFrame is None from {INPUT_FILE_PATH}")
            run_metrics["tasks"]["load_and_preprocess_data_duration_seconds"] = time.time() - task_start_time
            write_run_metrics(run_metrics, run_output_dir, run_id, pipeline_start_time, [], {}) # Pass empty list and empty dict
            return
    except Exception as e:
        logger.error(f"Error loading data in main: {e}", exc_info=True)
        run_metrics["errors_encountered"].append(f"Data loading exception: {str(e)}")
        run_metrics["tasks"]["load_and_preprocess_data_duration_seconds"] = time.time() - task_start_time
        write_run_metrics(run_metrics, run_output_dir, run_id, pipeline_start_time, [], {}) # Pass empty list and empty dict
        return
    run_metrics["tasks"]["load_and_preprocess_data_duration_seconds"] = time.time() - task_start_time

    if df is None:
        logger.error("DataFrame is None after loading attempt, cannot proceed.")
        return
    assert df is not None, "DataFrame loading failed."

    required_cols: Dict[str, Any] = {
        'ScrapingStatus': '',
        'RegexCandidateSnippets': lambda: [[] for _ in range(len(df))],
        'BestMatchedPhoneNumbers': lambda: [[] for _ in range(len(df))], 
        'OtherRelevantNumbers': lambda: [[] for _ in range(len(df))], 
        'ConfidenceScore': None, 
        'LLMExtractedNumbers': lambda: [[] for _ in range(len(df))], 
        'LLMContextPath': '',
        'Notes': '',
        'Top_Number_1': None,
        'Top_Type_1': None,
        'Top_SourceURL_1': None,
        'Top_Number_2': None,
        'Top_Type_2': None,
        'Top_SourceURL_2': None,
        'Top_Number_3': None,
        'Top_Type_3': None,
        'Top_SourceURL_3': None,
    }
    for col, default_val in required_cols.items():
        if col not in df.columns:
            if col == 'TargetCountryCodes' and col not in df.columns:
                 df[col] = pd.Series([[] for _ in range(len(df))], dtype=object)
            else:
                df[col] = default_val() if callable(default_val) else default_val
    
    if 'GivenPhoneNumber' not in df.columns:
        df['GivenPhoneNumber'] = None
    if 'Description' not in df.columns:
        df['Description'] = None
        df['Final_Row_Outcome_Reason'] = pd.Series([None] * len(df), dtype=object)
        df['Determined_Fault_Category'] = pd.Series([None] * len(df), dtype=object)

    # --- Pre-computation of Input Duplicate Counts ---
    logger.info("Starting pre-computation of input duplicate counts...")
    pre_comp_start_time = time.time()
    
    input_company_names_list: List[str] = []
    input_derived_canonical_urls_list: List[str] = [] # Store non-None, use placeholder for None

    # Use column names from AppConfig if available, otherwise default
    # These should have been mapped by load_and_preprocess_data
    company_name_col_key = 'CompanyName'
    url_col_key = 'GivenURL'

    if company_name_col_key not in df.columns:
        logger.error(f"Critical: Column '{company_name_col_key}' not found in DataFrame for duplicate pre-computation. This may lead to incorrect duplicate stats.")
    if url_col_key not in df.columns:
        logger.error(f"Critical: Column '{url_col_key}' not found in DataFrame for duplicate pre-computation. This may lead to incorrect duplicate stats.")

    for row_tuple in df.itertuples(index=False):
        company_name_val = str(getattr(row_tuple, company_name_col_key, "MISSING_COMPANY_NAME_INPUT")).strip()
        input_company_names_list.append(company_name_val)
        
        given_url_val = getattr(row_tuple, url_col_key, None)
        derived_input_canonical = get_input_canonical_url(given_url_val)
        input_derived_canonical_urls_list.append(derived_input_canonical if derived_input_canonical else "MISSING_OR_INVALID_URL_INPUT")

    company_name_counts = Counter(input_company_names_list)
    input_canonical_url_counts = Counter(input_derived_canonical_urls_list)

    run_metrics["data_processing_stats"]["input_unique_company_names"] = len(company_name_counts)
    run_metrics["data_processing_stats"]["input_unique_canonical_urls"] = len(input_canonical_url_counts)
    
    # Calculate number of items that are duplicates (appear > 1 times)
    num_company_names_with_duplicates = sum(1 for name, count in company_name_counts.items() if count > 1 and name != "MISSING_COMPANY_NAME_INPUT")
    num_urls_with_duplicates = sum(1 for url, count in input_canonical_url_counts.items() if count > 1 and url != "MISSING_OR_INVALID_URL_INPUT")
    run_metrics["data_processing_stats"]["input_company_names_with_duplicates_count"] = num_company_names_with_duplicates
    run_metrics["data_processing_stats"]["input_canonical_urls_with_duplicates_count"] = num_urls_with_duplicates

    # Calculate total rows affected by duplicates
    total_rows_with_dup_company = sum(count for name, count in company_name_counts.items() if count > 1 and name != "MISSING_COMPANY_NAME_INPUT")
    total_rows_with_dup_url = sum(count for url, count in input_canonical_url_counts.items() if count > 1 and url != "MISSING_OR_INVALID_URL_INPUT")
    run_metrics["data_processing_stats"]["input_rows_with_duplicate_company_name"] = total_rows_with_dup_company
    run_metrics["data_processing_stats"]["input_rows_with_duplicate_canonical_url"] = total_rows_with_dup_url
    
    # Calculate overall rows considered duplicates
    rows_considered_duplicates_overall = 0
    df['temp_input_canonical_url'] = input_derived_canonical_urls_list # Add as temp column for iteration
    df['temp_input_company_name'] = input_company_names_list # Add as temp column

    for _, row_data in df.iterrows():
        is_dup_company = company_name_counts[row_data['temp_input_company_name']] > 1 and row_data['temp_input_company_name'] != "MISSING_COMPANY_NAME_INPUT"
        is_dup_url = input_canonical_url_counts[row_data['temp_input_canonical_url']] > 1 and row_data['temp_input_canonical_url'] != "MISSING_OR_INVALID_URL_INPUT"
        if is_dup_company or is_dup_url:
            rows_considered_duplicates_overall += 1
    run_metrics["data_processing_stats"]["input_rows_considered_duplicates_overall"] = rows_considered_duplicates_overall
    
    df.drop(columns=['temp_input_canonical_url', 'temp_input_company_name'], inplace=True) # Clean up temp columns

    logger.info(f"Input duplicate pre-computation complete. Duration: {time.time() - pre_comp_start_time:.2f}s")
    logger.info(f"  Unique Company Names: {len(company_name_counts)}, Names with Duplicates: {num_company_names_with_duplicates}, Total Rows Affected by Company Duplicates: {total_rows_with_dup_company}")
    logger.info(f"  Unique Input Canonical URLs: {len(input_canonical_url_counts)}, URLs with Duplicates: {num_urls_with_duplicates}, Total Rows Affected by URL Duplicates: {total_rows_with_dup_url}")
    logger.info(f"  Total Input Rows Considered Duplicates (either Company or URL): {rows_considered_duplicates_overall}")
    run_metrics["tasks"]["pre_computation_duplicate_counts_duration_seconds"] = time.time() - pre_comp_start_time
    # --- End Pre-computation ---

    globally_processed_urls: Set[str] = set()
    all_flattened_rows: List[Dict[str, Any]] = []
    all_tertiary_rows: List[Dict[str, Any]] = []
    canonical_site_raw_llm_outputs: Dict[str, List[PhoneNumberLLMOutput]] = {}
    canonical_site_pathful_scraper_status: Dict[str, str] = {}
    input_to_canonical_map: Dict[str, Optional[str]] = {}
    canonical_site_regex_candidates_found_status: Dict[str, bool] = {} # New: Track if regex found candidates for a canonical
    canonical_site_llm_exception_details: Dict[str, str] = {} # New: Store specific LLM exception details
 
    pass1_loop_start_time = time.time()
    rows_processed_in_pass1 = 0
    rows_failed_in_pass1 = 0
    attrition_data_list: List[Dict[str, Any]] = [] # For Row Attrition Report
    row_level_failure_counts: Dict[str, int] = {} # Initialize counter for stage_of_failure

    with open(failure_log_csv_path, 'w', newline='', encoding='utf-8') as f_failure_log:
        failure_writer = csv.writer(f_failure_log)
        failure_writer.writerow([
            'log_timestamp', 'input_row_identifier', 'CompanyName', 'GivenURL',
            'stage_of_failure', 'error_reason', 'error_details',
            'Associated_Pathful_Canonical_URL' # New header
        ])

    # --- Data structures for new Canonical Domain Journey Report ---
    canonical_domain_journey_data: Dict[str, Dict[str, Any]] = {}
    # Helper dicts to aggregate data before putting into canonical_domain_journey_data
    true_base_to_input_row_ids: Dict[str, Set[Any]] = {}
    true_base_to_input_company_names: Dict[str, Set[str]] = {}
    true_base_to_input_given_urls: Dict[str, Set[str]] = {}
    true_base_to_pathful_urls_attempted: Dict[str, Set[str]] = {} # Stores unique pathful URLs processed under a true_base
    # true_base_scraper_status is already initialized later, will be used for this report too
    # canonical_site_regex_candidates_found_status is already initialized, will be aggregated
    # canonical_site_llm_exception_details is already initialized, will be aggregated
    # final_consolidated_data_by_true_base will be a key source

    # --- End Data structures for new Canonical Domain Journey Report ---

    for i, (index, row_series) in enumerate(df.iterrows()): # Corrected indentation
        rows_processed_in_pass1 += 1
        row: pd.Series = row_series
        company_name: str = str(row.get('CompanyName', f"Row_{index}"))
        given_url_original: Optional[str] = row.get('GivenURL')
        current_row_number_for_log: int = i + 1
        
        logger.info(f"[RowID: {index}, Company: {company_name}] --- Processing row {current_row_number_for_log}/{len(df)}: Original URL '{given_url_original}' ---")

        current_row_scraper_status: str = "Not_Run"
        final_canonical_entry_url: Optional[str] = None # Initialize for each row, before the try block
        # ... (other per-row initializations) ...
        
        given_url_original_str_key = str(given_url_original) if given_url_original is not None else "None_GivenURL_Input"
        processed_url = given_url_original

        try:
                if given_url_original and isinstance(given_url_original, str):
                    temp_url_stripped = given_url_original.strip()
                    parsed_obj = urlparse(temp_url_stripped)
                    current_scheme = parsed_obj.scheme
                    current_netloc = parsed_obj.netloc
                    current_path = parsed_obj.path
                    current_params = parsed_obj.params
                    current_query = parsed_obj.query
                    current_fragment = parsed_obj.fragment
                    if not current_scheme:
                        logger.info(f"[RowID: {index}, Company: {company_name}] URL '{temp_url_stripped}' is schemeless. Adding 'http://' and re-parsing.")
                        temp_for_reparse_schemeless = "http://" + temp_url_stripped
                        parsed_obj_schemed = urlparse(temp_for_reparse_schemeless)
                        current_scheme = parsed_obj_schemed.scheme 
                        current_netloc = parsed_obj_schemed.netloc
                        current_path = parsed_obj_schemed.path
                        current_params = parsed_obj_schemed.params 
                        current_query = parsed_obj_schemed.query   
                        current_fragment = parsed_obj_schemed.fragment 
                        logger.debug(f"[RowID: {index}, Company: {company_name}] After adding scheme: N='{current_netloc}', P='{current_path}'")
                    if " " in current_netloc:
                        logger.info(f"[RowID: {index}, Company: {company_name}] Spaces found in domain part '{current_netloc}'. Removing them.")
                        current_netloc = current_netloc.replace(" ", "")
                    current_path = quote(current_path, safe='/%')
                    current_query = quote(current_query, safe='=&/?+%')
                    current_fragment = quote(current_fragment, safe='/?#%')

                    # TLD Probing Logic
                    if current_netloc and not re.search(r'\.[a-zA-Z]{2,}$', current_netloc) and not current_netloc.endswith('.'):
                        is_ip_address = re.match(r"^\d{1,3}\.\d{1,3}\.\d{1,3}\.\d{1,3}$", current_netloc)
                        if current_netloc.lower() != 'localhost' and not is_ip_address:
                            logger.info(f"[RowID: {index}, Company: {company_name}] Input domain '{current_netloc}' appears to lack a TLD. Attempting TLD probing...")
                            successfully_probed_tld = False
                            probed_netloc_base = current_netloc # Keep original for logging if all probes fail
                            
                            for tld_to_try in app_config.url_probing_tlds:
                                candidate_domain_to_probe = f"{probed_netloc_base}.{tld_to_try}"
                                logger.debug(f"[RowID: {index}, Company: {company_name}] Probing TLD: Trying '{candidate_domain_to_probe}'")
                                try:
                                    socket.gethostbyname(candidate_domain_to_probe)
                                    current_netloc = candidate_domain_to_probe # Update current_netloc
                                    logger.info(f"[RowID: {index}, Company: {company_name}] TLD probe successful. Using '{current_netloc}' after trying '.{tld_to_try}'.")
                                    successfully_probed_tld = True
                                    break # Stop on first successful probe
                                except socket.gaierror:
                                    logger.debug(f"[RowID: {index}, Company: {company_name}] TLD probe DNS lookup failed for '{candidate_domain_to_probe}'.")
                                except Exception as sock_e: # Catch other potential socket errors
                                    logger.debug(f"[RowID: {index}, Company: {company_name}] TLD probe for '{candidate_domain_to_probe}' failed with unexpected socket error: {sock_e}")
                            
                            if not successfully_probed_tld:
                                logger.warning(f"[RowID: {index}, Company: {company_name}] TLD probing failed for base domain '{probed_netloc_base}'. Proceeding with '{current_netloc}' (which might be the un-suffixed original or last attempted).")
                                # current_netloc remains as it was before the loop if no probe succeeded (i.e. probed_netloc_base)
                                # Or, if a default append is desired here (e.g. .de as a last resort), it could be added.
                                # For now, we proceed with the netloc as is, and scraper will handle DNS failure.
                        
                    effective_path = current_path if current_path else ('/' if current_netloc else '')
                    
                    # Reconstruct the URL with potentially modified netloc
                    processed_url = urlparse('')._replace(
                        scheme=current_scheme, netloc=current_netloc, path=effective_path,
                        params=current_params, query=current_query, fragment=current_fragment
                    ).geturl()
                    
                    # Log the final decision for the URL to be scraped
                    if processed_url != given_url_original:
                        logger.info(f"[RowID: {index}, Company: {company_name}] URL: Original='{given_url_original}', Processed for Scraper='{processed_url}'")
                    else:
                        logger.info(f"[RowID: {index}, Company: {company_name}] URL: Using original='{given_url_original}' (no changes after preprocessing).")

                if not processed_url or not isinstance(processed_url, str) or not processed_url.startswith(('http://', 'https://')):
                    logger.warning(f"[RowID: {index}, Company: {company_name}] Skipping row {current_row_number_for_log} due to invalid or missing URL after all processing: '{processed_url}' (Original input was: '{given_url_original}')")
                    df.at[index, 'ScrapingStatus'] = 'InvalidURL'
                    current_row_scraper_status = 'InvalidURL'
                    df.at[index, 'VerificationStatus'] = 'Skipped_InvalidURL'
                    run_metrics["scraping_stats"]["scraping_failure_invalid_url"] += 1
                    log_row_failure(
                        failure_log_writer=failure_writer,
                        input_row_identifier=index,
                        company_name=company_name,
                        given_url=given_url_original,
                        stage_of_failure="URL_Validation_InvalidOrMissing",
                        error_reason=f"Invalid or missing URL after processing: {processed_url}",
                        log_timestamp=datetime.now().isoformat(),
                        error_details=json.dumps({"original_url": given_url_original, "processed_url": processed_url}),
                        # No specific pathful canonical URL yet for this type of input failure
                    )
                    stage_key = "URL_Validation_InvalidOrMissing"
                    row_level_failure_counts[stage_key] = row_level_failure_counts.get(stage_key, 0) + 1
                    rows_failed_in_pass1 +=1
                    continue
     
                scraped_pages_details: List[Tuple[str, str, str]]
                scraper_status: str
                # final_canonical_entry_url is now initialized at the start of the loop iteration
                
                run_metrics["scraping_stats"]["urls_processed_for_scraping"] += 1
                scrape_task_start_time = time.time()
                scraped_pages_details, scraper_status, final_canonical_entry_url = asyncio.run(
                    scrape_website(processed_url, run_output_dir, company_name, globally_processed_urls, index)
                )
                run_metrics["tasks"].setdefault("scrape_website_total_duration_seconds", 0)
                run_metrics["tasks"]["scrape_website_total_duration_seconds"] += (time.time() - scrape_task_start_time)

                df.at[index, 'ScrapingStatus'] = scraper_status
                true_base_domain_for_row = get_canonical_base_url(final_canonical_entry_url) if final_canonical_entry_url else None
                df.at[index, 'CanonicalEntryURL'] = true_base_domain_for_row
                current_row_scraper_status = scraper_status 
                given_url_original_str_key = str(given_url_original) if given_url_original is not None else "None_GivenURL_Input" 
                input_to_canonical_map[given_url_original_str_key] = true_base_domain_for_row

                # --- Start: Populate/Initialize structures for Canonical Domain Journey Report ---
                if true_base_domain_for_row:
                    if true_base_domain_for_row not in canonical_domain_journey_data:
                        canonical_domain_journey_data[true_base_domain_for_row] = {
                            "Input_Row_IDs": set(),
                            "Input_CompanyNames": set(),
                            "Input_GivenURLs": set(),
                            "Pathful_URLs_Attempted_List": set(),
                            "Overall_Scraper_Status_For_Domain": "Unknown", # Will be updated later
                            "Scraped_Pages_Details_Aggregated": Counter(), # page_type: count
                            "Total_Pages_Scraped_For_Domain": 0,
                            "Regex_Candidates_Found_For_Any_Pathful": False,
                            "LLM_Calls_Made_For_Domain": False,
                            "LLM_Total_Raw_Numbers_Extracted": 0,
                            "LLM_Total_Consolidated_Numbers_Found": 0, # Placeholder, updated after global consolidation
                            "LLM_Consolidated_Number_Types_Summary": Counter(), # Placeholder, updated after global consolidation
                            "LLM_Processing_Error_Encountered_For_Domain": False,
                            "LLM_Error_Messages_Aggregated": [],
                            "Final_Domain_Outcome_Reason": "Unknown", # Placeholder
                            "Primary_Fault_Category_For_Domain": "Unknown" # Placeholder
                        }
                    
                    canonical_domain_journey_data[true_base_domain_for_row]["Input_Row_IDs"].add(index)
                    canonical_domain_journey_data[true_base_domain_for_row]["Input_CompanyNames"].add(company_name)
                    if given_url_original:
                         canonical_domain_journey_data[true_base_domain_for_row]["Input_GivenURLs"].add(given_url_original)
                    if final_canonical_entry_url:
                        canonical_domain_journey_data[true_base_domain_for_row]["Pathful_URLs_Attempted_List"].add(final_canonical_entry_url)
                    
                    # Also populate the older helper dicts if they are still used elsewhere, or plan to deprecate them.
                    # For now, keeping them for compatibility during transition.
                    true_base_to_input_row_ids.setdefault(true_base_domain_for_row, set()).add(index)
                    true_base_to_input_company_names.setdefault(true_base_domain_for_row, set()).add(company_name)
                    if given_url_original:
                         true_base_to_input_given_urls.setdefault(true_base_domain_for_row, set()).add(given_url_original)
                    if final_canonical_entry_url:
                        true_base_to_pathful_urls_attempted.setdefault(true_base_domain_for_row, set()).add(final_canonical_entry_url)
                # --- End: Populate/Initialize structures ---

                logger.info(f"[RowID: {index}, Company: {company_name}] Row {current_row_number_for_log}: Scraper status: {current_row_scraper_status}, Pathful Canonical URL from Scraper: {final_canonical_entry_url}, True Base Domain: {true_base_domain_for_row}")

                if current_row_scraper_status == "Success" and final_canonical_entry_url:
                    if final_canonical_entry_url not in canonical_site_raw_llm_outputs: 
                        run_metrics["scraping_stats"]["new_canonical_sites_scraped"] += 1
                        run_metrics["regex_extraction_stats"]["sites_processed_for_regex"] += 1
                        regex_extraction_task_start_time = time.time()

                        logger.info(f"[RowID: {index}, Company: {company_name}] Processing new pathful canonical URL for LLM data collection: {final_canonical_entry_url} (from input {given_url_original})")
                        all_candidate_items_for_llm: List[Dict[str, str]] = []
                        if scraped_pages_details: 
                            run_metrics["scraping_stats"]["total_pages_scraped_overall"] += len(scraped_pages_details)
                            if final_canonical_entry_url not in run_metrics["scraping_stats"].get("processed_canonical_sites_for_success_count", set()):
                                run_metrics["scraping_stats"]["total_successful_canonical_scrapes"] += 1
                                run_metrics["scraping_stats"].setdefault("processed_canonical_sites_for_success_count", set()).add(final_canonical_entry_url)

                            target_codes_raw: Any = row.get('TargetCountryCodes', [])
                            target_codes_list_for_regex: List[str] = []
                            if isinstance(target_codes_raw, str) and target_codes_raw.startswith('[') and target_codes_raw.endswith(']'):
                                try:
                                    import ast
                                    parsed_eval = ast.literal_eval(target_codes_raw)
                                    if isinstance(parsed_eval, list):
                                        target_codes_list_for_regex = [str(item) for item in parsed_eval if isinstance(item, (str, int))]
                                except (ValueError, SyntaxError):
                                    logger.warning(f"[RowID: {index}, Company: {company_name}] Could not parse TargetCountryCodes string: {target_codes_raw}.")
                            elif isinstance(target_codes_raw, list):
                                target_codes_list_for_regex = [str(item) for item in target_codes_raw if isinstance(item, (str, int))]

                            for page_content_file, source_page_url, page_type in scraped_pages_details:
                                run_metrics["scraping_stats"]["pages_scraped_by_type"][page_type] = \
                                    run_metrics["scraping_stats"]["pages_scraped_by_type"].get(page_type, 0) + 1
                                
                                # --- Start: Aggregate page details for Canonical Domain Journey ---
                                if true_base_domain_for_row and true_base_domain_for_row in canonical_domain_journey_data:
                                    canonical_domain_journey_data[true_base_domain_for_row]["Scraped_Pages_Details_Aggregated"][page_type] += 1
                                    canonical_domain_journey_data[true_base_domain_for_row]["Total_Pages_Scraped_For_Domain"] += 1
                                # --- End: Aggregate page details ---
                                
                                if os.path.exists(page_content_file):
                                    try:
                                        with open(page_content_file, 'r', encoding='utf-8') as f_content:
                                            text_content = f_content.read()
                                        page_candidate_items: List[Dict[str, str]] = extract_numbers_with_snippets_from_text(
                                            text_content=text_content,
                                            source_url=source_page_url,
                                            original_input_company_name=company_name, 
                                            target_country_codes=target_codes_list_for_regex,
                                            snippet_window_chars=app_config.snippet_window_chars
                                        )
                                        all_candidate_items_for_llm.extend(page_candidate_items)
                                    except Exception as file_read_exc:
                                        logger.error(f"[RowID: {index}, Company: {company_name}] Error reading scraped page content {page_content_file} (canonical: {final_canonical_entry_url}): {file_read_exc}", exc_info=True)
                                        run_metrics["errors_encountered"].append(f"File read error for regex: {page_content_file}")
                                        log_row_failure(
                                            failure_log_writer=failure_writer,
                                            input_row_identifier=index,
                                            company_name=company_name,
                                            given_url=given_url_original,
                                            stage_of_failure="Regex_Extraction_FileReadError",
                                            error_reason="Error reading scraped content file",
                                            log_timestamp=datetime.now().isoformat(),
                                            error_details=json.dumps({
                                                "file_path": page_content_file,
                                                "canonical_url": final_canonical_entry_url, # This is a pathful canonical
                                                "exception": str(file_read_exc)
                                            }),
                                            associated_pathful_canonical_url=final_canonical_entry_url
                                        )
                                        stage_key = "Regex_Extraction_FileReadError"
                                        row_level_failure_counts[stage_key] = row_level_failure_counts.get(stage_key, 0) + 1
                                else:
                                    logger.warning(f"[RowID: {index}, Company: {company_name}] Scraped page content file not found: {page_content_file} (canonical: {final_canonical_entry_url})")
                            
                            run_metrics["tasks"].setdefault("regex_extraction_total_duration_seconds", 0)
                            run_metrics["tasks"]["regex_extraction_total_duration_seconds"] += (time.time() - regex_extraction_task_start_time)
                            if all_candidate_items_for_llm:
                                run_metrics["regex_extraction_stats"]["sites_with_regex_candidates"] += 1
                                run_metrics["regex_extraction_stats"]["total_regex_candidates_found"] += len(all_candidate_items_for_llm)
                                canonical_site_regex_candidates_found_status[final_canonical_entry_url] = True
                                # --- Start: Update Regex_Candidates_Found for Canonical Domain Journey ---
                                if true_base_domain_for_row and true_base_domain_for_row in canonical_domain_journey_data:
                                    canonical_domain_journey_data[true_base_domain_for_row]["Regex_Candidates_Found_For_Any_Pathful"] = True
                                # --- End: Update Regex_Candidates_Found ---
                            else:
                                canonical_site_regex_candidates_found_status[final_canonical_entry_url] = False
                                # No need to set Regex_Candidates_Found_For_Any_Pathful to False here, as it should remain True if any other pathful had candidates.
                            logger.info(f"[RowID: {index}, Company: {company_name}] Generated {len(all_candidate_items_for_llm)} candidate items for LLM for canonical URL {final_canonical_entry_url}. Regex candidates found: {canonical_site_regex_candidates_found_status[final_canonical_entry_url]}.")
     
                        if canonical_site_regex_candidates_found_status.get(final_canonical_entry_url, False): # Check if regex found candidates
                            run_metrics["llm_processing_stats"]["sites_processed_for_llm"] += 1
                            # --- Start: Update LLM_Calls_Made for Canonical Domain Journey ---
                            if true_base_domain_for_row and true_base_domain_for_row in canonical_domain_journey_data:
                                canonical_domain_journey_data[true_base_domain_for_row]["LLM_Calls_Made_For_Domain"] = True
                            # --- End: Update LLM_Calls_Made ---
                            llm_task_start_time = time.time()
                            try:
                                prompt_template_abs_path: str = app_config.llm_prompt_template_path
                                if not os.path.isabs(prompt_template_abs_path):
                                    project_root_dir_local = os.path.dirname(os.path.abspath(__file__))
                                    prompt_template_abs_path = os.path.join(project_root_dir_local, prompt_template_abs_path)

                                if not os.path.exists(prompt_template_abs_path):
                                    logger.error(f"[RowID: {index}, Company: {company_name}] LLM prompt template file not found at {prompt_template_abs_path}. Cannot process pathful canonical URL {final_canonical_entry_url}.")
                                    canonical_site_raw_llm_outputs[final_canonical_entry_url] = [] 
                                    canonical_site_pathful_scraper_status[final_canonical_entry_url] = "Error_LLM_PromptMissing"
                                    run_metrics["llm_processing_stats"]["llm_calls_failure_prompt_missing"] += 1
                                    run_metrics["errors_encountered"].append(f"LLM prompt template missing: {prompt_template_abs_path}")
                                    log_row_failure(
                                        failure_log_writer=failure_writer,
                                        input_row_identifier=index,
                                        company_name=company_name,
                                        given_url=given_url_original,
                                        stage_of_failure="LLM_Setup_PromptTemplateMissing",
                                        error_reason="LLM prompt template file not found",
                                        log_timestamp=datetime.now().isoformat(),
                                        error_details=json.dumps({
                                            "canonical_url": final_canonical_entry_url, # This is a pathful canonical
                                            "prompt_path": prompt_template_abs_path
                                        }),
                                        associated_pathful_canonical_url=final_canonical_entry_url
                                    )
                                    stage_key = "LLM_Setup_PromptTemplateMissing"
                                    row_level_failure_counts[stage_key] = row_level_failure_counts.get(stage_key, 0) + 1
                                else:
                                    safe_canonical_name_for_file = "".join(c if c.isalnum() else "_" for c in final_canonical_entry_url.replace("http://","").replace("https://",""))
                                    max_len_url_part = 100 
                                    if len(safe_canonical_name_for_file) > max_len_url_part:
                                        safe_canonical_name_for_file = safe_canonical_name_for_file[:max_len_url_part]
                                        logger.info(f"[RowID: {index}, Company: {company_name}] Truncated safe_canonical_name_for_file for {final_canonical_entry_url} to: {safe_canonical_name_for_file}")

                                    llm_input_filename = f"CANONICAL_{safe_canonical_name_for_file}_llm_input_data.json"
                                    llm_input_filepath = os.path.join(llm_context_dir, llm_input_filename)
                                    try:
                                        with open(llm_input_filepath, 'w', encoding='utf-8') as f_in: json.dump(all_candidate_items_for_llm, f_in, indent=2)
                                        logger.info(f"[RowID: {index}, Company: {company_name}] Saved LLM input data for {final_canonical_entry_url} to {llm_input_filepath}")
                                    except IOError as e: logger.error(f"[RowID: {index}, Company: {company_name}] IOError saving LLM input data for {final_canonical_entry_url}: {e}")

                                    llm_classified_outputs, llm_raw_response, token_stats = llm_extractor.extract_phone_numbers(
                                        candidate_items=all_candidate_items_for_llm,
                                        prompt_template_path=prompt_template_abs_path,
                                        llm_context_dir=llm_context_dir,
                                        file_identifier_prefix=f"CANONICAL_{safe_canonical_name_for_file}",
                                        triggering_input_row_id=index,
                                        triggering_company_name=company_name
                                    )
                                    canonical_site_raw_llm_outputs[final_canonical_entry_url] = llm_classified_outputs
                                    canonical_site_pathful_scraper_status[final_canonical_entry_url] = current_row_scraper_status
                                    run_metrics["llm_processing_stats"]["llm_calls_success"] += 1
                                    run_metrics["llm_processing_stats"]["total_llm_extracted_numbers_raw"] += len(llm_classified_outputs)
                                    # --- Start: Update LLM_Total_Raw_Numbers_Extracted for Canonical Domain Journey ---
                                    if true_base_domain_for_row and true_base_domain_for_row in canonical_domain_journey_data:
                                        canonical_domain_journey_data[true_base_domain_for_row]["LLM_Total_Raw_Numbers_Extracted"] += len(llm_classified_outputs)
                                    # --- End: Update LLM_Total_Raw_Numbers_Extracted ---

                                    if token_stats:
                                        run_metrics["llm_processing_stats"]["llm_successful_calls_with_token_data"] += 1
                                        run_metrics["llm_processing_stats"]["total_llm_prompt_tokens"] += token_stats.get("prompt_tokens", 0)
                                        run_metrics["llm_processing_stats"]["total_llm_completion_tokens"] += token_stats.get("completion_tokens", 0)
                                        run_metrics["llm_processing_stats"]["total_llm_tokens_overall"] += token_stats.get("total_tokens", 0)
                                        logger.info(f"[RowID: {index}, Company: {company_name}] LLM call for {final_canonical_entry_url} token usage: Prompt={token_stats.get('prompt_tokens',0)}, Completion={token_stats.get('completion_tokens',0)}, Total={token_stats.get('total_tokens',0)}")
                                    else:
                                        logger.warning(f"[RowID: {index}, Company: {company_name}] Token stats not available for LLM call related to {final_canonical_entry_url}")
                                    
                                    llm_raw_output_filename = f"CANONICAL_{safe_canonical_name_for_file}_llm_raw_output.json"
                                    llm_raw_output_filepath = os.path.join(llm_context_dir, llm_raw_output_filename)
                                    logger.info(f"[RowID: {index}, Company: {company_name}] Attempting to save LLM raw output. Path: '{llm_raw_output_filepath}', Length: {len(llm_raw_output_filepath)}")
                                    try:
                                        with open(llm_raw_output_filepath, 'w', encoding='utf-8') as f_llm_out:
                                            f_llm_out.write(llm_raw_response if isinstance(llm_raw_response, str) else json.dumps(llm_raw_response or {}, indent=2))
                                        logger.info(f"[RowID: {index}, Company: {company_name}] LLM classification for canonical {final_canonical_entry_url} complete. Raw output saved to {llm_raw_output_filepath}")
                                    except IOError as e:
                                        logger.error(f"[RowID: {index}, Company: {company_name}] IOError saving raw LLM output for {final_canonical_entry_url} to {llm_raw_output_filepath}: {e}")
                                        run_metrics["errors_encountered"].append(f"IOError saving LLM raw output: {llm_raw_output_filepath}")
                            except Exception as llm_exc:
                                logger.error(f"[RowID: {index}, Company: {company_name}] Error during LLM processing for pathful canonical {final_canonical_entry_url}: {llm_exc}", exc_info=True)
                                canonical_site_raw_llm_outputs[final_canonical_entry_url] = []
                                canonical_site_pathful_scraper_status[final_canonical_entry_url] = "Error_LLM_Processing"
                                # Capture the exception detail for the attrition report
                                exception_type_name = type(llm_exc).__name__
                                exception_message_str = str(llm_exc)
                                canonical_site_llm_exception_details[final_canonical_entry_url] = f"{exception_type_name}: {exception_message_str}"
                                # --- Start: Update LLM Error info for Canonical Domain Journey ---
                                if true_base_domain_for_row and true_base_domain_for_row in canonical_domain_journey_data:
                                    canonical_domain_journey_data[true_base_domain_for_row]["LLM_Processing_Error_Encountered_For_Domain"] = True
                                    canonical_domain_journey_data[true_base_domain_for_row]["LLM_Error_Messages_Aggregated"].append(f"PathfulURL ({final_canonical_entry_url}): {exception_type_name}: {exception_message_str}")
                                # --- End: Update LLM Error info ---
                                run_metrics["llm_processing_stats"]["llm_calls_failure_processing_error"] += 1
                                run_metrics["errors_encountered"].append(f"LLM processing error for {final_canonical_entry_url}: {str(llm_exc)}")
                                log_row_failure(
                                    failure_log_writer=failure_writer,
                                    input_row_identifier=index,
                                    company_name=company_name,
                                    given_url=given_url_original,
                                    stage_of_failure="LLM_Processing_GeneralError",
                                    error_reason="LLM processing error",
                                    log_timestamp=datetime.now().isoformat(),
                                    error_details=json.dumps({
                                        "canonical_url": final_canonical_entry_url, # This is a pathful canonical
                                        "exception_type": type(llm_exc).__name__,
                                        "exception_message": str(llm_exc)
                                    }),
                                    associated_pathful_canonical_url=final_canonical_entry_url
                                )
                                stage_key = "LLM_Processing_GeneralError"
                                row_level_failure_counts[stage_key] = row_level_failure_counts.get(stage_key, 0) + 1

                                run_metrics["tasks"].setdefault("llm_extraction_total_duration_seconds", 0)
                                run_metrics["tasks"]["llm_extraction_total_duration_seconds"] += (time.time() - llm_task_start_time)
                        else: # Corresponds to 'if canonical_site_regex_candidates_found_status.get(final_canonical_entry_url, False):'
                            logger.info(f"[RowID: {index}, Company: {company_name}] No regex candidate snippets for LLM from pathful canonical {final_canonical_entry_url}. Storing empty LLM result, LLM not called.")
                            canonical_site_raw_llm_outputs[final_canonical_entry_url] = [] # Ensure it's an empty list
                            canonical_site_pathful_scraper_status[final_canonical_entry_url] = current_row_scraper_status # Preserve scraper status
                            # Ensure llm_no_candidates_to_process is incremented if this canonical URL was new
                            # and would have been processed by LLM if candidates existed.
                            # This metric might need to be site-based rather than call-based if not already.
                            # For now, this correctly reflects that LLM was not called due to no candidates.
                            if final_canonical_entry_url not in run_metrics["llm_processing_stats"].get("sites_already_attempted_llm_or_skipped", set()):
                                run_metrics["llm_processing_stats"]["llm_no_candidates_to_process"] += 1
                                run_metrics["llm_processing_stats"].setdefault("sites_already_attempted_llm_or_skipped", set()).add(final_canonical_entry_url)
                    else: 
                        logger.info(f"[RowID: {index}, Company: {company_name}] Raw LLM data for pathful canonical URL {final_canonical_entry_url} already cached. Input row {given_url_original} maps to it.")
                
                elif current_row_scraper_status != "Success": 
                    logger.info(f"[RowID: {index}, Company: {company_name}] Row {current_row_number_for_log}: Scraper status '{current_row_scraper_status}'. No LLM processing for this input.")
                    if "Already_Processed" in current_row_scraper_status:
                        run_metrics["scraping_stats"]["scraping_failure_already_processed"] += 1
                    elif "InvalidURL" not in current_row_scraper_status : 
                        run_metrics["scraping_stats"]["scraping_failure_error"] += 1

                    if final_canonical_entry_url and final_canonical_entry_url not in canonical_site_pathful_scraper_status: 
                        canonical_site_pathful_scraper_status[final_canonical_entry_url] = current_row_scraper_status
                    df.at[index, 'Overall_VerificationStatus'] = f'Unverified_Scrape_{current_row_scraper_status}'
                    df.at[index, 'Original_Number_Status'] = f'Scrape_{current_row_scraper_status}' if row.get('NormalizedGivenPhoneNumber') else 'Original_Not_Provided'
                    log_row_failure(
                        failure_log_writer=failure_writer,
                        input_row_identifier=index,
                        company_name=company_name,
                        given_url=given_url_original,
                        stage_of_failure=f"Scraping_{current_row_scraper_status}",
                        error_reason=f"Scraper returned status: {current_row_scraper_status}",
                        log_timestamp=datetime.now().isoformat(),
                        error_details=json.dumps({
                            "pathful_canonical_url": final_canonical_entry_url, # This is a pathful canonical
                            "true_base_domain": true_base_domain_for_row
                        }),
                        associated_pathful_canonical_url=final_canonical_entry_url
                    )
                    stage_key = f"Scraping_{current_row_scraper_status}"
                    row_level_failure_counts[stage_key] = row_level_failure_counts.get(stage_key, 0) + 1
                    rows_failed_in_pass1 +=1
    
                if current_row_scraper_status == "Success":
                    run_metrics["scraping_stats"]["scraping_success"] += 1
                logger.info(f"[RowID: {index}, Company: {company_name}] Row {current_row_number_for_log}: Pass 1 processing complete. OriginalURL: {given_url_original_str_key}, CanonicalURL: {final_canonical_entry_url}, ScraperStatus: {current_row_scraper_status}")

        except Exception as e:
            logger.error(f"[RowID: {index}, Company: {company_name}] Error during Pass 1 processing for row {current_row_number_for_log}, Original URL {given_url_original_str_key}: {e}", exc_info=True)
            df.at[index, 'Overall_VerificationStatus'] = 'Error_Pass1_RowProcessing'
            current_scraper_status_for_df = df.at[index, 'ScrapingStatus']
            if current_scraper_status_for_df in ["Not_Run", "Success", None] or not current_scraper_status_for_df :
                 df.at[index, 'ScrapingStatus'] = f'PipelineError_{type(e).__name__}'
            run_metrics["errors_encountered"].append(f"Pass 1 row processing error for {company_name} (URL: {given_url_original_str_key}): {str(e)}")
            log_row_failure(
                failure_log_writer=failure_writer,
                input_row_identifier=index,
                company_name=company_name,
                given_url=given_url_original,
                stage_of_failure="RowProcessing_Pass1_UnhandledException",
                error_reason="Unhandled exception during Pass 1 row processing",
                log_timestamp=datetime.now().isoformat(),
                error_details=json.dumps({
                    "exception_type": type(e).__name__,
                    "exception_message": str(e)
                }),
                # Use final_canonical_entry_url (which was initialized at loop start) if available,
                # otherwise what's in df (if populated by an earlier, successful scrape for this input row), or None.
                associated_pathful_canonical_url=final_canonical_entry_url if final_canonical_entry_url else (df.at[index, 'CanonicalEntryURL'] if 'CanonicalEntryURL' in df.columns and pd.notna(df.at[index, 'CanonicalEntryURL']) else None)
            )
            stage_key = "RowProcessing_Pass1_UnhandledException"
            row_level_failure_counts[stage_key] = row_level_failure_counts.get(stage_key, 0) + 1
            rows_failed_in_pass1 +=1
            logger.error(
                f"[RowID: {index}, Company: {company_name}] Row {current_row_number_for_log} errored in Pass 1. "
                f"ScraperStatus='{df.at[index, 'ScrapingStatus']}', "
                f"OverallVerificationStatus='{df.at[index, 'Overall_VerificationStatus']}'"
            )
            for col_prefix in ['Primary_', 'Secondary_']:
                for suffix in ['Number_1', 'Type_1', 'SourceURL_1', 'Number_2', 'Type_2', 'SourceURL_2']:
                    col_name = f"{col_prefix}{suffix}"
                    if col_name in df.columns:
                        df.at[index, col_name] = None
            if 'Original_Number_Status' in df.columns:
                df.at[index, 'Original_Number_Status'] = 'Error_Pass1_RowProcessing'
        
    run_metrics["tasks"]["pass1_main_loop_duration_seconds"] = time.time() - pass1_loop_start_time
    run_metrics["data_processing_stats"]["rows_successfully_processed_pass1"] = rows_processed_in_pass1 - rows_failed_in_pass1
    run_metrics["data_processing_stats"]["rows_failed_pass1"] = rows_failed_in_pass1
    run_metrics["data_processing_stats"]["row_level_failure_summary"] = row_level_failure_counts # Store the collected counts
    logger.info(f"Pass 1 (Scraping and Raw LLM Data Collection) complete. Processed {rows_processed_in_pass1} input rows.")
    logger.info(f"Unique pathful canonical sites for which raw LLM data was collected: {len(canonical_site_raw_llm_outputs)}")
    logger.debug(f"Pathful canonical site raw LLM data cache keys: {list(canonical_site_raw_llm_outputs.keys())}")
    logger.debug(f"Pathful canonical site scraper status cache: {list(canonical_site_pathful_scraper_status.keys())}")
    logger.debug(f"Input to True Base Domain map entries: {len(input_to_canonical_map)}")
    
    run_metrics["scraping_stats"]["total_urls_fetched_by_scraper"] = run_metrics["scraping_stats"]["total_pages_scraped_overall"]

    if "processed_canonical_sites_for_success_count" in run_metrics["scraping_stats"]:
        del run_metrics["scraping_stats"]["processed_canonical_sites_for_success_count"]
 
    global_consolidation_start_time = time.time()
    logger.info("Starting Global Consolidation of LLM data by True Base Domain...")
    final_consolidated_data_by_true_base: Dict[str, Optional[CompanyContactDetails]] = {}
    true_base_to_pathful_map: Dict[str, List[str]] = {}
    true_base_to_input_company_names: Dict[str, Set[str]] = {}
    true_base_scraper_status: Dict[str, str] = {}


    for pathful_url_key, raw_llm_list in canonical_site_raw_llm_outputs.items():
        true_base = get_canonical_base_url(pathful_url_key)
        if not true_base:
            logger.warning(f"Could not get true_base_domain for pathful_url_key '{pathful_url_key}' during global consolidation. Skipping.")
            continue
        
        if true_base not in true_base_to_pathful_map:
            true_base_to_pathful_map[true_base] = []
            true_base_to_input_company_names[true_base] = set()
            true_base_scraper_status[true_base] = "Unknown" 
        
        true_base_to_pathful_map[true_base].append(pathful_url_key)
        
        current_pathful_status = canonical_site_pathful_scraper_status.get(pathful_url_key, "Unknown")
        
        # Update true_base_scraper_status (existing logic)
        if true_base_scraper_status[true_base] == "Unknown" or \
           (current_pathful_status == "Success" and true_base_scraper_status[true_base] != "Success") or \
           ("Error" not in current_pathful_status and "Error" in true_base_scraper_status[true_base]):
            true_base_scraper_status[true_base] = current_pathful_status

        # --- Start: Update Overall_Scraper_Status_For_Domain in canonical_domain_journey_data ---
        if true_base in canonical_domain_journey_data:
            # Use the same logic as for true_base_scraper_status
            current_domain_overall_status = canonical_domain_journey_data[true_base].get("Overall_Scraper_Status_For_Domain", "Unknown")
            if current_domain_overall_status == "Unknown" or \
               (current_pathful_status == "Success" and current_domain_overall_status != "Success") or \
               ("Error" not in current_pathful_status and "Error" in current_domain_overall_status):
                canonical_domain_journey_data[true_base]["Overall_Scraper_Status_For_Domain"] = current_pathful_status
        # --- End: Update Overall_Scraper_Status_For_Domain ---


    if 'CanonicalEntryURL' in df.columns and 'CompanyName' in df.columns:
        for true_base_domain_key in true_base_to_pathful_map.keys():
            mask = df['CanonicalEntryURL'].notna() & (df['CanonicalEntryURL'] == true_base_domain_key)
            matching_companies = df.loc[mask, 'CompanyName'].dropna().astype(str).unique()
            if len(matching_companies) > 0:
                 true_base_to_input_company_names[true_base_domain_key].update(matching_companies)
            else: 
                 first_pathful_for_base = true_base_to_pathful_map[true_base_domain_key][0]
                 logger.warning(f"No matching company names found in df for true_base_domain '{true_base_domain_key}'. Company name in report might be from first pathful trigger.")


    for true_base_domain, list_of_pathful_urls in true_base_to_pathful_map.items():
        all_llm_results_for_this_true_base: List[PhoneNumberLLMOutput] = []
        for pathful_url_item in list_of_pathful_urls:
            all_llm_results_for_this_true_base.extend(canonical_site_raw_llm_outputs.get(pathful_url_item, []))
        
        representative_company_name_for_consolidation = "Unknown"
        if true_base_to_input_company_names.get(true_base_domain):
            representative_company_name_for_consolidation = sorted(list(true_base_to_input_company_names[true_base_domain]))[0]
        elif list_of_pathful_urls: 
            pass


        final_consolidated_data_by_true_base[true_base_domain] = process_and_consolidate_contact_data(
            llm_results=all_llm_results_for_this_true_base,
            company_name_from_input=representative_company_name_for_consolidation,
            initial_given_url=true_base_domain
        )
        # --- Start: Populate LLM consolidated numbers info in canonical_domain_journey_data ---
        if true_base_domain in canonical_domain_journey_data and final_consolidated_data_by_true_base[true_base_domain]:
            consolidated_details = final_consolidated_data_by_true_base[true_base_domain]
            if consolidated_details and consolidated_details.consolidated_numbers:
                canonical_domain_journey_data[true_base_domain]["LLM_Total_Consolidated_Numbers_Found"] = len(consolidated_details.consolidated_numbers)
                for cn_item in consolidated_details.consolidated_numbers:
                    for source_detail in cn_item.sources:
                        if source_detail.type: # Ensure type is not None
                             canonical_domain_journey_data[true_base_domain]["LLM_Consolidated_Number_Types_Summary"][source_detail.type] += 1
        # --- End: Populate LLM consolidated numbers info ---
    logger.info(f"Global Consolidation complete. {len(final_consolidated_data_by_true_base)} true base domains processed.")
    run_metrics["tasks"]["global_consolidation_duration_seconds"] = time.time() - global_consolidation_start_time
    run_metrics["data_processing_stats"]["unique_true_base_domains_consolidated"] = len(final_consolidated_data_by_true_base)

    # --- Determine Final Outcome and Fault Category for each Canonical Domain ---
    logger.info("Determining final outcomes for each canonical domain...")
    for domain_key, journey_data_entry in canonical_domain_journey_data.items():
        final_domain_reason, final_domain_fault = _determine_final_domain_outcome_and_fault(
            true_base_domain=domain_key,
            domain_journey_entry=journey_data_entry,
            true_base_scraper_status_map=true_base_scraper_status, # This was populated during global consolidation
            true_base_to_pathful_map=true_base_to_pathful_map, # This was populated during global consolidation
            canonical_site_pathful_scraper_status=canonical_site_pathful_scraper_status, # Populated in Pass 1
            # For canonical_site_regex_candidates_found_status, we use the aggregated value from the journey data
            canonical_site_regex_candidates_found_status=journey_data_entry.get("Regex_Candidates_Found_For_Any_Pathful", False),
            final_consolidated_data=final_consolidated_data_by_true_base.get(domain_key) # Result from process_and_consolidate
        )
        canonical_domain_journey_data[domain_key]["Final_Domain_Outcome_Reason"] = final_domain_reason
        canonical_domain_journey_data[domain_key]["Primary_Fault_Category_For_Domain"] = final_domain_fault
        if final_domain_reason != "Contact_Successfully_Extracted_For_Domain":
            logger.info(f"Domain '{domain_key}': Outcome='{final_domain_reason}', Fault='{final_domain_fault}'")
    logger.info("Final domain outcome determination complete.")
    # --- End Final Outcome Determination for Canonical Domains ---
 
 
    detailed_columns_order = [
        'CompanyName', 'Number', 'LLM_Type', 'LLM_Classification', 
        'LLM_Source_URL', 'ScrapingStatus', 'TargetCountryCodes', 'RunID'
    ]

    summary_columns_order = [
        'CompanyName', 'GivenURL', 'GivenPhoneNumber', 'Original_Number_Status',
        'Top_Number_1', 'Top_Type_1', 'Description', 'ScrapingStatus_Canonical', 
        'CanonicalEntryURL', 'Top_Number_1', 'Top_Type_1', 'Top_Number_2', 
        'Top_Type_2', 'Top_Number_3', 'Top_Type_3', 'Top_SourceURL_1', 
        'Top_SourceURL_2', 'Top_SourceURL_3', 'TargetCountryCodes', 'RunID'
    ]

    tertiary_report_columns_order = [
        'CompanyName', 'GivenURL', 'CanonicalEntryURL', 'ScrapingStatus', 
        'PhoneNumber_1', 'PhoneNumber_2', 'PhoneNumber_3', 
        'SourceURL_1', 'SourceURL_2', 'SourceURL_3'
    ]

    logger.info("Starting Pass 2: Building Detailed Flattened and Summary Reports...")
    pass2_reports_start_time = time.time()
 
    classification_precedence = { 
        'Primary': 1, 'Secondary': 2, 'Support': 3,
        'Low Relevance': 4, 'Non-Business': 5, None: 99
    }

    for index, original_row_data in df.iterrows(): 
        company_name_pass2 = str(original_row_data.get('CompanyName', f"Row_{index}"))
        given_url_pass2 = original_row_data.get('GivenURL')
        canonical_url_pass2 = original_row_data.get('CanonicalEntryURL')
        scraper_status_pass2 = original_row_data.get('ScrapingStatus')

        company_contact_details_pass2: Optional[CompanyContactDetails] = None
        if canonical_url_pass2 and canonical_url_pass2 in final_consolidated_data_by_true_base: 
            company_contact_details_pass2 = final_consolidated_data_by_true_base[canonical_url_pass2]
        
        scraper_status_for_true_base_detailed = true_base_scraper_status.get(str(canonical_url_pass2), "Unknown") if canonical_url_pass2 else "Unknown_NoTrueBase"

        if scraper_status_for_true_base_detailed == "Success" and company_contact_details_pass2 and company_contact_details_pass2.consolidated_numbers:
            for consolidated_number_item in company_contact_details_pass2.consolidated_numbers:
                aggregated_types = []
                aggregated_source_urls = []
                seen_types_for_number = set() 
                
                for source_detail in consolidated_number_item.sources:
                    type_with_path = f"{source_detail.type} (from {source_detail.source_path})"
                    if source_detail.type not in seen_types_for_number: 
                        aggregated_types.append(source_detail.type)
                        seen_types_for_number.add(source_detail.type)
                    if source_detail.original_full_url not in aggregated_source_urls: 
                         aggregated_source_urls.append(source_detail.original_full_url)

                llm_type_str = ", ".join(aggregated_types) if aggregated_types else consolidated_number_item.sources[0].type if consolidated_number_item.sources else "Unknown"
                llm_source_url_str = ", ".join(aggregated_source_urls) if aggregated_source_urls else consolidated_number_item.sources[0].original_full_url if consolidated_number_item.sources else "N/A"

                new_flattened_row: Dict[str, Any] = {
                    'CompanyName': company_name_pass2,
                    'Number': consolidated_number_item.number,
                    'LLM_Type': llm_type_str, 
                    'LLM_Classification': consolidated_number_item.classification, 
                    'LLM_Source_URL': llm_source_url_str, 
                    'ScrapingStatus': scraper_status_for_true_base_detailed, 
                    'TargetCountryCodes': original_row_data.get('TargetCountryCodes'),
                    'RunID': run_id 
                }
                all_flattened_rows.append(new_flattened_row)
        

    logger.info("Starting Aggregation for Top_Contacts_Report...")
    top_contacts_aggregation_map: Dict[str, Dict[str, Any]] = {}

    if 'CanonicalEntryURL' not in df.columns:
        logger.error("'CanonicalEntryURL' column is missing from DataFrame. This is critical for Top_Contacts_Report aggregation. Initializing to None.")
        df['CanonicalEntryURL'] = None
    if 'CompanyName' not in df.columns:
        logger.error("'CompanyName' column is missing from DataFrame. This is critical for Top_Contacts_Report aggregation. Initializing to 'Unknown_Input_Company'.")
        df['CompanyName'] = "Unknown_Input_Company"
    if 'GivenURL' not in df.columns:
        logger.error("'GivenURL' column is missing from DataFrame. This is critical for Top_Contacts_Report aggregation. Initializing to 'Unknown_Input_GivenURL'.")
        df['GivenURL'] = "Unknown_Input_GivenURL"

    for true_base_domain_key_agg, company_contact_details_object in final_consolidated_data_by_true_base.items():
        if company_contact_details_object is None:
            logger.warning(f"Skipping true_base_domain '{true_base_domain_key_agg}' for Top_Contacts_Report aggregation as its CompanyContactDetails is None.")
            continue

        matching_input_rows = df[df['CanonicalEntryURL'].astype(str) == str(true_base_domain_key_agg)]

        unique_original_company_names: Set[str]
        unique_original_given_urls: Set[str]

        if matching_input_rows.empty:
            logger.warning(f"No input rows found in df mapping to true_base_domain '{true_base_domain_key_agg}'. "
                           f"Using company name ('{company_contact_details_object.company_name}') from CompanyContactDetails and its original_input_urls as fallback.")
            unique_original_company_names = {str(company_contact_details_object.company_name)} if company_contact_details_object.company_name else {"Unknown_Company"}
            unique_original_given_urls = set(map(str, company_contact_details_object.original_input_urls))
        else:
            unique_original_company_names = set(matching_input_rows['CompanyName'].dropna().astype(str))
            unique_original_given_urls = set(matching_input_rows['GivenURL'].dropna().astype(str))
            if not unique_original_company_names:
                 unique_original_company_names = {str(company_contact_details_object.company_name)} if company_contact_details_object.company_name else {"Unknown_Company"}
            if not unique_original_given_urls:
                 unique_original_given_urls = set(map(str, company_contact_details_object.original_input_urls))

        report_company_name = f"{true_base_domain_key_agg} - {' - '.join(sorted(list(unique_original_company_names)))}"
        report_given_urls = ", ".join(sorted(list(unique_original_given_urls)))
        
        top_contacts_aggregation_map[true_base_domain_key_agg] = {
            "report_company_name": report_company_name,
            "report_given_urls": report_given_urls,
            "canonical_entry_url": true_base_domain_key_agg, 
            "scraper_status": true_base_scraper_status.get(true_base_domain_key_agg, "Unknown"), 
            "contact_details": company_contact_details_object,
            "all_input_companies_for_canonical": sorted(list(unique_original_company_names))
        }
    logger.info(f"Aggregation for Top_Contacts_Report complete. Found {len(top_contacts_aggregation_map)} unique canonical URLs to report on.")

    logger.info("Building Top_Contacts_Report (formerly Tertiary) from aggregated data...")
    all_tertiary_rows.clear()

    for aggregated_entry in top_contacts_aggregation_map.values():
        company_contact_details_for_report = aggregated_entry["contact_details"]
        
        new_tertiary_row: Dict[str, Any] = {
            'CompanyName': aggregated_entry["report_company_name"],
            'GivenURL': aggregated_entry["report_given_urls"],
            'CanonicalEntryURL': aggregated_entry["canonical_entry_url"],
            'ScrapingStatus': aggregated_entry["scraper_status"],
            'PhoneNumber_1': None, 'PhoneNumber_2': None, 'PhoneNumber_3': None,
            'SourceURL_1': None, 'SourceURL_2': None, 'SourceURL_3': None
        }

        if company_contact_details_for_report and company_contact_details_for_report.consolidated_numbers:
            eligible_numbers_for_report: List[ConsolidatedPhoneNumber] = []
            for cn_item in company_contact_details_for_report.consolidated_numbers:
                source_types = {s.type for s in cn_item.sources if s.type}
                if cn_item.classification != 'Non-Business' and \
                   not EXCLUDED_TYPES_FOR_TOP_CONTACTS_REPORT.intersection(source_types):
                    eligible_numbers_for_report.append(cn_item)
                else:
                    excluded_reasons = []
                    if cn_item.classification == 'Non-Business':
                        excluded_reasons.append("classification is Non-Business")
                    intersecting_types = EXCLUDED_TYPES_FOR_TOP_CONTACTS_REPORT.intersection(source_types)
                    if intersecting_types:
                        excluded_reasons.append(f"excluded types: {intersecting_types}")
                    logger.debug(f"Excluding number {cn_item.number} from Top_Contacts_Report for company '{aggregated_entry['report_company_name']}' due to: {'; '.join(excluded_reasons)}")

            for i, consolidated_number_item in enumerate(eligible_numbers_for_report[:3]): 
                phone_num_key = f'PhoneNumber_{i+1}'
                source_url_key = f'SourceURL_{i+1}'
                
                number_str = consolidated_number_item.number
                types_str = ", ".join(sorted(list(set(s.type for s in consolidated_number_item.sources))))
                
                companies_for_this_number = sorted(list(set(
                    s.original_input_company_name
                    for s in consolidated_number_item.sources
                    if s.original_input_company_name
                )))
                companies_for_this_number_str = ", ".join(companies_for_this_number) if companies_for_this_number else "UnknownCompany" 

                new_tertiary_row[phone_num_key] = f"{number_str} ({types_str}) [{companies_for_this_number_str}]"
                new_tertiary_row[source_url_key] = ", ".join(sorted(list(set(s.original_full_url for s in consolidated_number_item.sources))))
        
        if new_tertiary_row['PhoneNumber_1'] or new_tertiary_row['PhoneNumber_2'] or new_tertiary_row['PhoneNumber_3']:
            all_tertiary_rows.append(new_tertiary_row)
        else:
            logger.info(f"Skipping row for canonical URL '{aggregated_entry['canonical_entry_url']}' (Company: '{aggregated_entry['report_company_name']}') in Top_Contacts_Report as it has no eligible phone numbers after filtering.")
            
    logger.info(f"Finished building Top_Contacts_Report. {len(all_tertiary_rows)} rows created.")


    for index, row_summary in df.iterrows():
        # Preserve original values needed by the helper or for other logic
        given_url_original_for_attrition = str(row_summary.get('GivenURL')) if row_summary.get('GivenURL') is not None else ""
        company_name_for_attrition = str(row_summary.get('CompanyName', f"Row_{index}"))
        canonical_url_summary = row_summary.get('CanonicalEntryURL') # This is the true_base_domain

        # Snapshot of relevant df statuses for this row before they might be overwritten
        df_status_snapshot_for_helper = {
            'ScrapingStatus': row_summary.get('ScrapingStatus'), # Initial scrape status for the input URL
            # Add other df statuses if _determine_final_row_outcome_and_fault needs them
        }

        company_contact_details_summary: Optional[CompanyContactDetails] = None
        if canonical_url_summary and canonical_url_summary in final_consolidated_data_by_true_base:
            company_contact_details_summary = final_consolidated_data_by_true_base[canonical_url_summary]

        unique_sorted_consolidated_numbers: List[ConsolidatedPhoneNumber] = []
        if company_contact_details_summary:
            unique_sorted_consolidated_numbers = company_contact_details_summary.consolidated_numbers

        # --- Populate Top Numbers (existing logic, kept for now) ---
        if len(unique_sorted_consolidated_numbers) > 0:
            top_item_1 = unique_sorted_consolidated_numbers[0]
            df.at[index, 'Top_Number_1'] = top_item_1.number
            df.at[index, 'Top_Type_1'] = ", ".join(list(set(s.type for s in top_item_1.sources)))
            df.at[index, 'Top_SourceURL_1'] = ", ".join(list(set(s.original_full_url for s in top_item_1.sources)))
        if len(unique_sorted_consolidated_numbers) > 1:
            top_item_2 = unique_sorted_consolidated_numbers[1]
            df.at[index, 'Top_Number_2'] = top_item_2.number
            df.at[index, 'Top_Type_2'] = ", ".join(list(set(s.type for s in top_item_2.sources)))
            df.at[index, 'Top_SourceURL_2'] = ", ".join(list(set(s.original_full_url for s in top_item_2.sources)))
        if len(unique_sorted_consolidated_numbers) > 2:
            top_item_3 = unique_sorted_consolidated_numbers[2]
            df.at[index, 'Top_Number_3'] = top_item_3.number
            df.at[index, 'Top_Type_3'] = ", ".join(list(set(s.type for s in top_item_3.sources)))
            df.at[index, 'Top_SourceURL_3'] = ", ".join(list(set(s.original_full_url for s in top_item_3.sources)))
        # --- End Populate Top Numbers ---

        # --- Determine Final Row Outcome Reason and Fault Category ---
        final_reason, fault_category = _determine_final_row_outcome_and_fault(
            index=index,
            row_summary=row_summary,
            df_status_snapshot=df_status_snapshot_for_helper,
            company_contact_details_summary=company_contact_details_summary,
            unique_sorted_consolidated_numbers=unique_sorted_consolidated_numbers,
            canonical_url_summary=canonical_url_summary,
            true_base_scraper_status_map=true_base_scraper_status,
            true_base_to_pathful_map=true_base_to_pathful_map,
            canonical_site_pathful_scraper_status=canonical_site_pathful_scraper_status,
            canonical_site_raw_llm_outputs=canonical_site_raw_llm_outputs,
            canonical_site_regex_candidates_found_status=canonical_site_regex_candidates_found_status, # Pass new dict
            canonical_site_llm_exception_details=canonical_site_llm_exception_details # Pass new dict
        )
        df.at[index, 'Final_Row_Outcome_Reason'] = final_reason
        df.at[index, 'Determined_Fault_Category'] = fault_category

        # --- Log the determined outcome for the input row ---
        if final_reason != "Contact_Successfully_Extracted":
            logger.info(
                f"[RowID: {index}, Company: {company_name_for_attrition}] Input row outcome: {final_reason} (Fault: {fault_category}). GivenURL: '{given_url_original_for_attrition}'"
            )
            # --- Populate Attrition Data List ---
            # Get pre-computed duplicate counts for this row
            current_input_company_name_for_counts = str(row_summary.get(company_name_col_key, "MISSING_COMPANY_NAME_INPUT")).strip()
            current_input_given_url_for_counts = row_summary.get(url_col_key)
            current_derived_input_canonical_for_counts = get_input_canonical_url(current_input_given_url_for_counts)
            if not current_derived_input_canonical_for_counts:
                current_derived_input_canonical_for_counts = "MISSING_OR_INVALID_URL_INPUT"

            input_company_total_count = company_name_counts.get(current_input_company_name_for_counts, 0)
            input_url_total_count = input_canonical_url_counts.get(current_derived_input_canonical_for_counts, 0)

            is_dup_company_name = input_company_total_count > 1 and current_input_company_name_for_counts != "MISSING_COMPANY_NAME_INPUT"
            is_dup_input_url = input_url_total_count > 1 and current_derived_input_canonical_for_counts != "MISSING_OR_INVALID_URL_INPUT"
            is_overall_input_duplicate = is_dup_company_name or is_dup_input_url

            attrition_data_list.append({
                "InputRowID": index,
                "CompanyName": company_name_for_attrition,
                "GivenURL": given_url_original_for_attrition,
                "Final_Row_Outcome_Reason": final_reason,
                "Determined_Fault_Category": fault_category,
                "Relevant_Canonical_URLs": canonical_url_summary if canonical_url_summary else "N/A",
                "LLM_Error_Detail_Summary": canonical_site_llm_exception_details.get(str(canonical_url_summary), "") if canonical_url_summary and fault_category == "LLM Issue" and "Error" in final_reason else "",
                "Timestamp_Of_Determination": datetime.now().isoformat(),
                "Input_CompanyName_Total_Count": input_company_total_count,
                "Input_CanonicalURL_Total_Count": input_url_total_count,
                "Is_Input_CompanyName_Duplicate": "Yes" if is_dup_company_name else "No",
                "Is_Input_CanonicalURL_Duplicate": "Yes" if is_dup_input_url else "No",
                "Is_Input_Row_Considered_Duplicate": "Yes" if is_overall_input_duplicate else "No"
            })
        else:
             logger.info(
                f"[RowID: {index}, Company: {company_name_for_attrition}] Input row outcome: {final_reason}. GivenURL: '{given_url_original_for_attrition}'"
            )


        # --- Original_Number_Status logic (kept for now, might be simplified later) ---
        original_norm_phone_summary = row_summary.get('NormalizedGivenPhoneNumber')
        if original_norm_phone_summary and original_norm_phone_summary != "InvalidFormat":
            found_original_in_top_llm = False
            for top_num_item in unique_sorted_consolidated_numbers[:3]:
                if top_num_item.number == original_norm_phone_summary:
                    found_original_in_top_llm = True
                    break
            if found_original_in_top_llm:
                df.at[index, 'Original_Number_Status'] = 'Verified'
            elif unique_sorted_consolidated_numbers:
                df.at[index, 'Original_Number_Status'] = 'Corrected'
            # The following conditions are now largely covered by Final_Row_Outcome_Reason
            # We might simplify this or make it dependent on Final_Row_Outcome_Reason
            elif final_reason == "LLM_Output_NoNumbersFound_AllAttempts" or final_reason == "LLM_Output_NumbersFound_NoneRelevant_AllAttempts":
                 df.at[index, 'Original_Number_Status'] = 'LLM_OutputEmpty_Or_NoRelevant_For_Canonical' # Keep old status for compatibility if needed
            elif final_reason.startswith("LLM_NoInput") or final_reason.startswith("LLM_Processing_Error") or final_reason.startswith("ScrapingFailed_Canonical") or final_reason == "Unknown_NoCanonicalURLDetermined":
                 df.at[index, 'Original_Number_Status'] = 'LLM_Not_Run_Or_NoOutput_For_Canonical' # Keep old status
            else: # Fallback if original phone was provided but no LLM match and not covered above
                 df.at[index, 'Original_Number_Status'] = 'No Relevant Match Found by LLM'

        elif original_norm_phone_summary == "InvalidFormat":
            df.at[index, 'Original_Number_Status'] = 'Original_InvalidFormat'
        else:
            df.at[index, 'Original_Number_Status'] = 'Original_Not_Provided'
        # --- End Original_Number_Status logic ---

        # --- Overall_VerificationStatus logic (kept for now, review for simplification later) ---
        # This logic might be redundant or can be derived from Final_Row_Outcome_Reason
        overall_status = "Unverified"
        scraper_status_for_true_base_domain_summary = true_base_scraper_status.get(str(canonical_url_summary), "Unknown") if canonical_url_summary else "Unknown_NoTrueBase"

        if final_reason == "Contact_Successfully_Extracted":
            overall_status = "Verified_LLM_Match_Found"
        elif final_reason.startswith("ScrapingFailed_Canonical") or final_reason == "Scraping_AllAttemptsFailed_Network" or final_reason == "Scraping_AllAttemptsFailed_AccessDenied" or final_reason == "Scraping_ContentNotFound_AllAttempts":
            overall_status = f"Unverified_Scrape_{scraper_status_for_true_base_domain_summary}"
        elif final_reason == "LLM_Output_NumbersFound_NoneRelevant_AllAttempts" or final_reason == "LLM_Output_NoNumbersFound_AllAttempts":
            overall_status = "Unverified_LLM_NoRelevantNumbers"
        elif final_reason == "LLM_Processing_Error_AllAttempts": # Covers prompt missing too if helper sets it
            overall_status = "Error_LLM_Processing_For_Canonical"
        # Add more mappings from final_reason to overall_status if needed

        original_input_url_for_map = str(row_summary.get('GivenURL')) if row_summary.get('GivenURL') is not None else "None_GivenURL_Input"
        normalized_original_input_base = get_canonical_base_url(original_input_url_for_map, log_level_for_non_domain_input=logging.INFO) if original_input_url_for_map != "None_GivenURL_Input" else None
        
        if canonical_url_summary and normalized_original_input_base and normalized_original_input_base != canonical_url_summary:
            if overall_status != "Verified_LLM_Match_Found" or not overall_status.startswith("RedirectedTo"): # Avoid double prefix
                 overall_status = f"RedirectedTo[{canonical_url_summary}]_" + overall_status
        
        df.at[index, 'Overall_VerificationStatus'] = overall_status
        df.at[index, 'ScrapingStatus_Canonical'] = scraper_status_for_true_base_domain_summary
        df.at[index, 'LLM_Processing_Status_Canonical'] = "Processed" if company_contact_details_summary is not None and unique_sorted_consolidated_numbers else scraper_status_for_true_base_domain_summary

    if all_flattened_rows:
        df_detailed_flattened = pd.DataFrame(all_flattened_rows)
        
        classification_sort_order = ['Primary', 'Secondary', 'Support', 'Low Relevance', 'Non-Business']
        df_detailed_flattened['LLM_Classification_Sort'] = pd.Categorical(
            df_detailed_flattened['LLM_Classification'],
            categories=classification_sort_order,
            ordered=True
        )
        df_detailed_flattened = df_detailed_flattened.sort_values(
            by=['CompanyName', 'LLM_Classification_Sort', 'Number'],
            na_position='last' 
        ).drop(columns=['LLM_Classification_Sort'])
        
        for col in detailed_columns_order:
            if col not in df_detailed_flattened.columns:
                df_detailed_flattened[col] = None 
        
        df_detailed_export = df_detailed_flattened[detailed_columns_order].copy()

        detailed_output_filename = f"All_LLM_Extractions_Report_{run_id}.xlsx" 
        detailed_output_excel_path = os.path.join(run_output_dir, detailed_output_filename)
        try:
            logger.info(f"Attempting to save detailed report to: {detailed_output_excel_path}")
            with pd.ExcelWriter(detailed_output_excel_path, engine='openpyxl') as writer:
                df_detailed_export.to_excel(writer, index=False, sheet_name='Detailed_Phone_Data')
                worksheet_detailed = writer.sheets['Detailed_Phone_Data']
                for col_idx, col_name in enumerate(df_detailed_export.columns):
                    series_data = df_detailed_export.iloc[:, col_idx]
                    if series_data.empty:
                        max_val_len = 0
                    else:
                        lengths = series_data.astype(str).map(len)
                        max_val_len = lengths.max() if not lengths.empty else 0
                    
                    column_header_len = len(str(col_name))
                    adjusted_width = max(max_val_len, column_header_len) + 2
                    worksheet_detailed.column_dimensions[get_column_letter(col_idx + 1)].width = adjusted_width
            logger.info(f"Detailed report saved successfully to {detailed_output_excel_path}")
        except Exception as e_detailed:
            logger.error(f"Error saving detailed report to {detailed_output_excel_path}: {e_detailed}", exc_info=True)
    else:
        logger.info("No data for detailed flattened report. Skipping file creation.")
    run_metrics["report_generation_stats"]["detailed_report_rows"] = len(all_flattened_rows)
 
    unique_summary_cols_needed = list(dict.fromkeys(summary_columns_order)) 

    for col_name in unique_summary_cols_needed:
        if col_name not in df.columns:
            if col_name in ['Original_Number_Status', 'Overall_VerificationStatus',
                            'CanonicalEntryURL', 'ScrapingStatus_Canonical', 'LLM_Processing_Status_Canonical',
                            'Top_Number_1', 'Top_Type_1', 'Top_SourceURL_1',
                            'Top_Number_2', 'Top_Type_2', 'Top_SourceURL_2',
                            'Top_Number_3', 'Top_Type_3', 'Top_SourceURL_3']:
                df[col_name] = None 
                logger.warning(f"Summary report column '{col_name}' was not found in DataFrame and was initialized to None. Check population logic.")
            elif col_name == 'RunID':
                df[col_name] = run_id 
            elif col_name not in ['CompanyName', 'GivenURL', 'GivenPhoneNumber', 'Description', 'TargetCountryCodes']:
                 logger.error(f"Unexpected summary column '{col_name}' missing and not covered by specific initialization. Defaulting to None.")
                 df[col_name] = None


    df_summary_export = df[summary_columns_order].copy()
    
    summary_output_filename = app_config.output_excel_file_name_template.format(run_id=run_id)
    summary_output_excel_path = os.path.join(run_output_dir, summary_output_filename)
    try:
        logger.info(f"Attempting to save summary report to: {summary_output_excel_path}")
        with pd.ExcelWriter(summary_output_excel_path, engine='openpyxl') as writer:
            df_summary_export.to_excel(writer, index=False, sheet_name='Phone_Validation_Summary')
            worksheet_summary = writer.sheets['Phone_Validation_Summary']
            for col_idx, col_name in enumerate(df_summary_export.columns):
                series_data = df_summary_export.iloc[:, col_idx]
                if series_data.empty:
                    max_val_len = 0
                else:
                    lengths = series_data.astype(str).map(len)
                    max_val_len = lengths.max() if not lengths.empty else 0
                
                column_header_len = len(str(col_name))
                adjusted_width = max(max_val_len, column_header_len) + 2
                worksheet_summary.column_dimensions[get_column_letter(col_idx + 1)].width = adjusted_width
        logger.info(f"Summary report saved successfully to {summary_output_excel_path}")
    except Exception as e_summary:
        logger.error(f"Error saving summary report to {summary_output_excel_path}: {e_summary}", exc_info=True)
        run_metrics["errors_encountered"].append(f"Error saving summary report: {str(e_summary)}")
    run_metrics["report_generation_stats"]["summary_report_rows"] = len(df_summary_export) if 'df_summary_export' in locals() else 0
 
    tertiary_output_filename = app_config.tertiary_report_file_name_template.format(run_id=run_id) 
    tertiary_output_excel_path = os.path.join(run_output_dir, tertiary_output_filename)

    if all_tertiary_rows:
        df_tertiary_report = pd.DataFrame(all_tertiary_rows)
        
        for col_t in tertiary_report_columns_order:
            if col_t not in df_tertiary_report.columns:
                df_tertiary_report[col_t] = None 
        
        df_tertiary_export = df_tertiary_report[tertiary_report_columns_order].copy()

        try:
            logger.info(f"Attempting to save tertiary report ('Final Contacts.xlsx') to: {tertiary_output_excel_path}")
            with pd.ExcelWriter(tertiary_output_excel_path, engine='openpyxl') as writer_t:
                df_tertiary_export.to_excel(writer_t, index=False, sheet_name='Contact_Focused_Report')
                worksheet_tertiary = writer_t.sheets['Contact_Focused_Report']
                for col_idx, col_name in enumerate(df_tertiary_export.columns):
                    series_data = df_tertiary_export.iloc[:, col_idx]
                    if series_data.empty:
                        max_val_len = 0
                    else:
                        lengths = series_data.astype(str).map(len)
                        max_val_len = lengths.max() if not lengths.empty else 0
                    
                    column_header_len = len(str(col_name))
                    adjusted_width = max(max_val_len, column_header_len) + 2
                    worksheet_tertiary.column_dimensions[get_column_letter(col_idx + 1)].width = adjusted_width
            logger.info(f"Tertiary report saved successfully to {tertiary_output_excel_path}")
        except Exception as e_tertiary:
            logger.error(f"Error saving tertiary report to {tertiary_output_excel_path}: {e_tertiary}", exc_info=True)
            run_metrics["errors_encountered"].append(f"Error saving tertiary report: {str(e_tertiary)}")
    else:
        logger.info("No data for tertiary report. Skipping file creation.")
    run_metrics["report_generation_stats"]["tertiary_report_rows"] = len(all_tertiary_rows)
    run_metrics["tasks"]["pass2_report_generation_duration_seconds"] = time.time() - pass2_reports_start_time
    
    # Write Canonical Domain Summary Report
    logger.info("Attempting to write Canonical Domain Summary Report...")
    canonical_domain_summary_rows_written = write_canonical_domain_summary_report(
        run_id=run_id,
        domain_journey_data=canonical_domain_journey_data,
        output_dir=run_output_dir
    )
    run_metrics["report_generation_stats"]["canonical_domain_summary_rows"] = canonical_domain_summary_rows_written
    logger.info(f"Canonical Domain Summary Report: {canonical_domain_summary_rows_written} rows written.")

    # Write Row Attrition Report before final metrics
    num_attrition_rows = write_row_attrition_report(run_id, attrition_data_list, run_output_dir, canonical_domain_journey_data, input_to_canonical_map)
    # Potentially add num_attrition_rows to run_metrics if needed for summary in run_metrics.md
    run_metrics["data_processing_stats"]["rows_in_attrition_report"] = num_attrition_rows
 
    write_run_metrics(run_metrics, run_output_dir, run_id, pipeline_start_time, attrition_data_list, canonical_domain_journey_data)

    logger.info("Attempting to generate 'Final Processed Contacts' report...")
    if os.path.exists(tertiary_output_excel_path):
        try:
            generate_processed_contacts_report(
                final_contacts_file_path=tertiary_output_excel_path, 
                config=app_config,
                run_id=run_id
            )
        except Exception as e_processed_report:
            logger.error(f"Error generating 'Final Processed Contacts' report: {e_processed_report}", exc_info=True)
    else:
        logger.warning(f"'Final Contacts.xlsx' not found at {tertiary_output_excel_path}. Skipping 'Final Processed Contacts' report generation.")

    logger.info(f"Pipeline run {run_id} finished.")
    logger.info(f"Total pipeline duration: {run_metrics['total_duration_seconds']:.2f} seconds.")
    logger.info(f"Run metrics file created at: {os.path.join(run_output_dir, f'run_metrics_{run_id}.md')}")


def write_row_attrition_report(run_id: str, attrition_data: List[Dict[str, Any]], output_dir: str, canonical_domain_journey_data: Dict[str, Dict[str, Any]], input_to_canonical_map: Dict[str, Optional[str]]) -> int:
    """Writes the collected row attrition data to an Excel file with auto-width columns."""
    if not attrition_data:
        logger.info("No data for row attrition report. Skipping file creation.")
        return 0

    report_filename = f"row_attrition_report_{run_id}.xlsx"
    report_path = os.path.join(output_dir, report_filename)
    report_df = pd.DataFrame(attrition_data)
    
    # Ensure consistent column order, including LLM_Error_Detail_Summary and new duplicate fields
    # Update attrition data with new fields before creating DataFrame
    for row_data in attrition_data:
        given_url = row_data.get("GivenURL")
        row_data["Derived_Input_CanonicalURL"] = get_input_canonical_url(given_url)
        
        # 'Relevant_Canonical_URLs' in attrition_data is expected to be the true_base_domain
        # for the input row if one was determined.
        final_processed_domain = row_data.get("Relevant_Canonical_URLs")
        row_data["Final_Processed_Canonical_Domain"] = final_processed_domain if pd.notna(final_processed_domain) and final_processed_domain != "N/A" else None

        link_to_outcome = None
        if final_processed_domain and pd.notna(final_processed_domain) and final_processed_domain != "N/A":
            if final_processed_domain in canonical_domain_journey_data:
                # The "link" is the domain key itself, which can be used for lookups/hyperlinks
                link_to_outcome = final_processed_domain
            else:
                # This case might occur if an input row failed very early,
                # and its 'Relevant_Canonical_URLs' was populated from df['CanonicalEntryURL']
                # but that domain didn't make it into canonical_domain_journey_data (e.g. if it was a duplicate canonical not processed further)
                # Or if the input_to_canonical_map has a mapping but that canonical wasn't processed.
                # Try to find it via input_to_canonical_map if direct lookup fails
                input_url_key = str(given_url) if given_url is not None else "None_GivenURL_Input"
                mapped_canonical = input_to_canonical_map.get(input_url_key)
                if mapped_canonical and mapped_canonical in canonical_domain_journey_data:
                    link_to_outcome = mapped_canonical
                else:
                    logger.debug(f"AttritionReport: Could not find domain '{final_processed_domain}' (from GivenURL: {given_url}) in canonical_domain_journey_data for linking.")
        row_data["Link_To_Canonical_Domain_Outcome"] = link_to_outcome

    report_df = pd.DataFrame(attrition_data)

    columns_order = [
        "InputRowID", "CompanyName", "GivenURL",
        "Derived_Input_CanonicalURL", # New
        "Final_Processed_Canonical_Domain", # New
        "Link_To_Canonical_Domain_Outcome", # New
        "Final_Row_Outcome_Reason", "Determined_Fault_Category",
        "Relevant_Canonical_URLs", # This is the old field, kept for reference, might be same as Final_Processed_Canonical_Domain
        "LLM_Error_Detail_Summary",
        "Input_CompanyName_Total_Count",
        "Input_CanonicalURL_Total_Count",
        "Is_Input_CompanyName_Duplicate",
        "Is_Input_CanonicalURL_Duplicate",
        "Is_Input_Row_Considered_Duplicate",
        "Timestamp_Of_Determination"
    ]
    
    # Add any missing columns to the DataFrame with None, and reorder
    for col in columns_order:
        if col not in report_df.columns:
            report_df[col] = None # Initialize missing columns with None
    report_df = report_df[columns_order]

    try:
        with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
            report_df.to_excel(writer, index=False, sheet_name='Attrition_Report')
            worksheet = writer.sheets['Attrition_Report']
            for col_idx, col_name in enumerate(report_df.columns):
                series_data = report_df.iloc[:, col_idx]
                if series_data.empty:
                    max_val_len = 0
                else:
                    # Ensure all data is string for length calculation
                    lengths = series_data.astype(str).map(len)
                    max_val_len = lengths.max() if not lengths.empty else 0
                
                column_header_len = len(str(col_name))
                adjusted_width = max(max_val_len, column_header_len) + 2 # Add a little padding
                worksheet.column_dimensions[get_column_letter(col_idx + 1)].width = adjusted_width
        
        logger.info(f"Row attrition report successfully saved to {report_path}")
        return len(report_df)
    except Exception as e: # Changed from IOError to general Exception for broader catch
        logger.error(f"Failed to write row attrition report to {report_path}: {e}", exc_info=True)
        return 0


def write_run_metrics(metrics: Dict[str, Any], output_dir: str, run_id: str, pipeline_start_time: float, attrition_data_list_for_metrics: List[Dict[str, Any]], canonical_domain_journey_data: Dict[str, Dict[str, Any]]) -> None:
    """Writes the collected run metrics to a Markdown file."""
    metrics["total_duration_seconds"] = time.time() - pipeline_start_time
    metrics_file_path = os.path.join(output_dir, f"run_metrics_{run_id}.md")

    try:
        with open(metrics_file_path, 'w', encoding='utf-8') as f:
            f.write(f"# Pipeline Run Metrics: {run_id}\n\n")
            f.write(f"**Run ID:** {metrics.get('run_id', 'N/A')}\n")
            f.write(f"**Total Run Duration:** {metrics.get('total_duration_seconds', 0):.2f} seconds\n")
            f.write(f"**Pipeline Start Time:** {datetime.fromtimestamp(pipeline_start_time).strftime('%Y-%m-%d %H:%M:%S')}\n")
            f.write(f"**Pipeline End Time:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n")

            f.write("## Task Durations (seconds):\n")
            if metrics.get("tasks"):
                for task_name, duration in metrics["tasks"].items():
                    f.write(f"- **{task_name.replace('_', ' ').title()}:** {duration:.2f}\n")
            else:
                f.write("- No task durations recorded.\n")
            f.write("\n")

            f.write("### Average Task Durations (per relevant item):\n")

            tasks_data = metrics.get("tasks", {})
            scraping_stats_data = metrics.get("scraping_stats", {})
            regex_stats_data = metrics.get("regex_extraction_stats", {})
            llm_stats_data = metrics.get("llm_processing_stats", {})
            data_proc_stats_data = metrics.get("data_processing_stats", {})

            # Average Scrape Duration
            total_scrape_duration = tasks_data.get("scrape_website_total_duration_seconds", 0)
            new_canonical_sites_scraped = scraping_stats_data.get("new_canonical_sites_scraped", 0)
            if new_canonical_sites_scraped > 0:
                avg_scrape_duration = total_scrape_duration / new_canonical_sites_scraped
                f.write(f"- **Average Scrape Website Duration (per New Canonical Site Scraped):** {avg_scrape_duration:.2f} seconds\n")
            else:
                f.write("- Average Scrape Website Duration (per New Canonical Site Scraped): N/A (No new canonical sites scraped)\n")

            # Average Regex Extraction Duration
            total_regex_duration = tasks_data.get("regex_extraction_total_duration_seconds", 0)
            sites_processed_for_regex = regex_stats_data.get("sites_processed_for_regex", 0)
            if sites_processed_for_regex > 0:
                avg_regex_duration = total_regex_duration / sites_processed_for_regex
                f.write(f"- **Average Regex Extraction Duration (per Site Processed for Regex):** {avg_regex_duration:.2f} seconds\n")
            else:
                f.write("- Average Regex Extraction Duration (per Site Processed for Regex): N/A (No sites processed for regex)\n")

            # Average LLM Extraction Duration
            total_llm_duration = tasks_data.get("llm_extraction_total_duration_seconds", 0)
            sites_processed_for_llm = llm_stats_data.get("sites_processed_for_llm", 0)
            if sites_processed_for_llm > 0:
                avg_llm_duration = total_llm_duration / sites_processed_for_llm
                f.write(f"- **Average LLM Extraction Duration (per Site Processed for LLM):** {avg_llm_duration:.2f} seconds\n")
            else:
                f.write("- Average LLM Extraction Duration (per Site Processed for LLM): N/A (No sites processed for LLM)\n")

            # Average Pass 1 Main Loop Duration
            total_pass1_duration = tasks_data.get("pass1_main_loop_duration_seconds", 0)
            input_rows_count = data_proc_stats_data.get("input_rows_count", 0)
            if input_rows_count > 0:
                avg_pass1_duration = total_pass1_duration / input_rows_count
                f.write(f"- **Average Pass 1 Main Loop Duration (per Input Row):** {avg_pass1_duration:.2f} seconds\n")
            else:
                f.write("- Average Pass 1 Main Loop Duration (per Input Row): N/A (No input rows)\n")
            
            f.write("\n") # Add a newline before the next section

            f.write("## Data Processing Statistics:\n")
            stats = metrics.get("data_processing_stats", {})
            f.write(f"- **Input Rows Processed (Initial Load):** {stats.get('input_rows_count', 0)}\n")
            f.write(f"- **Rows Successfully Processed (Pass 1):** {stats.get('rows_successfully_processed_pass1', 0)}\n")
            f.write(f"- **Rows Failed During Processing (Pass 1):** {stats.get('rows_failed_pass1', 0)} (Input rows that did not complete Pass 1 successfully due to errors such as invalid URL, scraping failure, or critical processing exceptions for that row, preventing LLM processing or final data consolidation for that specific input.)\n")
            f.write(f"- **Unique True Base Domains Consolidated:** {stats.get('unique_true_base_domains_consolidated', 0)}\n")
            # Removed extra \n\n to place new section directly after

            f.write("\n## Input Data Duplicate Analysis:\n")
            dp_stats = metrics.get("data_processing_stats", {})
            f.write(f"- **Total Input Rows Analyzed for Duplicates:** {dp_stats.get('input_rows_count', 0)}\n")
            f.write(f"- **Unique Input CompanyNames Found:** {dp_stats.get('input_unique_company_names', 0)}\n")
            f.write(f"- **Input CompanyNames Appearing More Than Once:** {dp_stats.get('input_company_names_with_duplicates_count', 0)}\n")
            f.write(f"- **Total Input Rows with a Duplicate CompanyName:** {dp_stats.get('input_rows_with_duplicate_company_name', 0)}\n")
            f.write(f"- **Unique Input Canonical URLs Found:** {dp_stats.get('input_unique_canonical_urls', 0)}\n")
            f.write(f"- **Input Canonical URLs Appearing More Than Once:** {dp_stats.get('input_canonical_urls_with_duplicates_count', 0)}\n")
            f.write(f"- **Total Input Rows with a Duplicate Input Canonical URL:** {dp_stats.get('input_rows_with_duplicate_canonical_url', 0)}\n")
            f.write(f"- **Total Input Rows Considered Duplicates (either CompanyName or URL):** {dp_stats.get('input_rows_considered_duplicates_overall', 0)}\n\n")

            # Analyze duplicates within the attrition list
            attrition_input_company_duplicates = 0
            attrition_input_url_duplicates = 0
            attrition_overall_input_duplicates = 0
            if attrition_data_list_for_metrics:
                for attrition_row in attrition_data_list_for_metrics:
                    if attrition_row.get("Is_Input_CompanyName_Duplicate") == "Yes":
                        attrition_input_company_duplicates += 1
                    if attrition_row.get("Is_Input_CanonicalURL_Duplicate") == "Yes":
                        attrition_input_url_duplicates += 1
                    if attrition_row.get("Is_Input_Row_Considered_Duplicate") == "Yes":
                        attrition_overall_input_duplicates += 1
            
            f.write("### Input Duplicates within Attrition Report:\n")
            total_attrition_rows = len(attrition_data_list_for_metrics)
            f.write(f"- **Total Rows in Attrition Report:** {total_attrition_rows}\n")
            f.write(f"- **Attrition Rows with Original Input CompanyName Duplicate:** {attrition_input_company_duplicates}\n")
            f.write(f"- **Attrition Rows with Original Input Canonical URL Duplicate:** {attrition_input_url_duplicates}\n")
            f.write(f"- **Attrition Rows Considered Overall Original Input Duplicates:** {attrition_overall_input_duplicates}\n\n")


            f.write("## Scraping Statistics:\n")
            stats = metrics.get("scraping_stats", {})
            f.write(f"- **URLs Processed for Scraping:** {stats.get('urls_processed_for_scraping', 0)}\n")
            f.write(f"- **New Canonical Sites Scraped:** {stats.get('new_canonical_sites_scraped', 0)}\n")
            f.write(f"- **Scraping Successes:** {stats.get('scraping_success', 0)}\n")
            f.write(f"- **Scraping Failures (Invalid URL):** {stats.get('scraping_failure_invalid_url', 0)}\n")
            f.write(f"- **Scraping Failures (Already Processed):** {stats.get('scraping_failure_already_processed', 0)}\n")
            f.write(f"- **Scraping Failures (Other Errors):** {stats.get('scraping_failure_error', 0)}\n")
            f.write(f"- **Total Pages Scraped Overall:** {stats.get('total_pages_scraped_overall', 0)}\n")
            f.write(f"- **Total Unique URLs Successfully Fetched:** {stats.get('total_urls_fetched_by_scraper', 0)}\n") 
            f.write(f"- **Total Successfully Scraped Canonical Sites:** {stats.get('total_successful_canonical_scrapes', 0)}\n") 

            if stats.get('total_successful_canonical_scrapes', 0) > 0:
                avg_pages_per_site = stats.get('total_pages_scraped_overall', 0) / stats.get('total_successful_canonical_scrapes', 1) 
                f.write(f"- **Average Pages Scraped per Successfully Scraped Canonical Site:** {avg_pages_per_site:.2f}\n")
            else:
                f.write("- Average Pages Scraped per Successfully Scraped Canonical Site: N/A (No successful canonical scrapes)\n")

            f.write("- **Pages Scraped by Type:**\n")
            pages_by_type = stats.get("pages_scraped_by_type", {})
            if pages_by_type:
                for page_type, count in sorted(pages_by_type.items()): 
                    f.write(f"  - *{page_type.replace('_', ' ').title()}:* {count}\n")
            else:
                f.write("  - No page type data recorded.\n")
            f.write("\n")

            f.write("## Regex Extraction Statistics:\n")
            stats = metrics.get("regex_extraction_stats", {})
            f.write(f"- **Canonical Sites Processed for Regex:** {stats.get('sites_processed_for_regex', 0)}\n")
            f.write(f"- **Canonical Sites with Regex Candidates Found:** {stats.get('sites_with_regex_candidates', 0)}\n")
            f.write(f"- **Total Regex Candidates Found:** {stats.get('total_regex_candidates_found', 0)}\n\n")

            f.write("## LLM Processing Statistics:\n")
            stats = metrics.get("llm_processing_stats", {})
            f.write(f"- **Canonical Sites Sent for LLM Processing:** {stats.get('sites_processed_for_llm', 0)}\n")
            f.write(f"- **LLM Calls Successful:** {stats.get('llm_calls_success', 0)}\n")
            f.write(f"- **LLM Calls Failed (Prompt Missing):** {stats.get('llm_calls_failure_prompt_missing', 0)}\n")
            f.write(f"- **LLM Calls Failed (Processing Error):** {stats.get('llm_calls_failure_processing_error', 0)}\n")
            f.write(f"- **Canonical Sites with No Regex Candidates (Skipped LLM):** {stats.get('llm_no_candidates_to_process', 0)}\n")
            f.write(f"- **Total LLM Extracted Phone Number Objects (Raw):** {stats.get('total_llm_extracted_numbers_raw', 0)}\n")
            f.write(f"- **LLM Successful Calls with Token Data:** {stats.get('llm_successful_calls_with_token_data', 0)}\n")
            f.write(f"- **Total LLM Prompt Tokens:** {stats.get('total_llm_prompt_tokens', 0)}\n")
            f.write(f"- **Total LLM Completion Tokens:** {stats.get('total_llm_completion_tokens', 0)}\n")
            f.write(f"- **Total LLM Tokens Overall:** {stats.get('total_llm_tokens_overall', 0)}\n")

            successful_calls_for_avg = stats.get('llm_successful_calls_with_token_data', 0)
            if successful_calls_for_avg > 0:
                avg_prompt_tokens = stats.get('total_llm_prompt_tokens', 0) / successful_calls_for_avg
                avg_completion_tokens = stats.get('total_llm_completion_tokens', 0) / successful_calls_for_avg
                avg_total_tokens = stats.get('total_llm_tokens_overall', 0) / successful_calls_for_avg
                f.write(f"- **Average Prompt Tokens per Successful Call:** {avg_prompt_tokens:.2f}\n")
                f.write(f"- **Average Completion Tokens per Successful Call:** {avg_completion_tokens:.2f}\n")
                f.write(f"- **Average Total Tokens per Successful Call:** {avg_total_tokens:.2f}\n")
            else:
                f.write("- Average token counts not available (no successful calls with token data).\n")
            f.write("\n")

            f.write("## Report Generation Statistics:\n")
            stats = metrics.get("report_generation_stats", {})
            f.write(f"- **Detailed Report Rows Created:** {stats.get('detailed_report_rows', 0)}\n")
            f.write(f"- **Summary Report Rows Created:** {stats.get('summary_report_rows', 0)}\n")
            f.write(f"- **Tertiary Report Rows Created:** {stats.get('tertiary_report_rows', 0)}\n")
            f.write(f"- **Canonical Domain Summary Rows Created:** {stats.get('canonical_domain_summary_rows', 0)}\n\n")

            f.write("## Canonical Domain Processing Summary:\n")
            if canonical_domain_journey_data:
                total_canonical_domains = len(canonical_domain_journey_data)
                f.write(f"- **Total Unique Canonical Domains Processed:** {total_canonical_domains}\n")

                outcome_counts: Dict[str, int] = Counter()
                fault_counts: Dict[str, int] = Counter()

                for domain_data in canonical_domain_journey_data.values():
                    outcome = domain_data.get("Final_Domain_Outcome_Reason", "Unknown_Outcome")
                    fault = domain_data.get("Primary_Fault_Category_For_Domain", "Unknown_Fault")
                    outcome_counts[outcome] += 1
                    if fault != "N/A": # Only count actual faults
                        fault_counts[fault] += 1
                
                f.write("### Outcomes for Canonical Domains:\n")
                if outcome_counts:
                    for outcome, count in sorted(outcome_counts.items()):
                        f.write(f"  - **{outcome.replace('_', ' ').title()}:** {count}\n")
                else:
                    f.write("  - No outcome data recorded for canonical domains.\n")
                f.write("\n")

                f.write("### Primary Fault Categories for Canonical Domains (where applicable):\n")
                if fault_counts:
                    for fault, count in sorted(fault_counts.items()):
                        f.write(f"  - **{fault.replace('_', ' ').title()}:** {count}\n")
                else:
                    f.write("  - No fault data recorded for canonical domains or all succeeded.\n")
                f.write("\n")
            else:
                f.write("- No canonical domain journey data available to summarize.\n\n")

            f.write("## Summary of Row-Level Failures (from `failed_rows_{run_id}.csv`):\n")
            row_failures_summary = metrics.get("data_processing_stats", {}).get("row_level_failure_summary", {})

            if row_failures_summary:
                grouped_failures: Dict[str, Dict[str, Any]] = {
                    "Scraping": {"total": 0, "details": {}},
                    "LLM": {"total": 0, "details": {}},
                    "URL Validation": {"total": 0, "details": {}},
                    "Regex Extraction": {"total": 0, "details": {}},
                    "Row Processing": {"total": 0, "details": {}},
                    "Other": {"total": 0, "details": {}}
                }
                failure_category_map: Dict[str, str] = {
                    "Scraping_": "Scraping",
                    "LLM_": "LLM",
                    "URL_Validation_": "URL Validation",
                    "Regex_Extraction_": "Regex Extraction",
                    "RowProcessing_": "Row Processing"
                }

                for stage, count in sorted(row_failures_summary.items()):
                    matched_category = False
                    for prefix, category_name in failure_category_map.items():
                        if stage.startswith(prefix):
                            grouped_failures[category_name]["total"] += count
                            grouped_failures[category_name]["details"][stage] = count
                            matched_category = True
                            break
                    if not matched_category:
                        grouped_failures["Other"]["total"] += count
                        grouped_failures["Other"]["details"][stage] = count

                for category_name, data in grouped_failures.items():
                    if data["total"] > 0:
                        f.write(f"- **Total {category_name} Failures:** {data['total']}\n")
                        for stage, count in sorted(data["details"].items()):
                            f.write(f"  - *{stage.replace('_', ' ').title()}:* {count}\n")
                f.write("\n")
            else:
                f.write("- No row-level failures recorded with specific stages.\n")
            f.write("\n")
            
            f.write("## Global Pipeline Errors:\n") # Renamed section
            if metrics.get("errors_encountered"):
                for error_msg in metrics["errors_encountered"]:
                    f.write(f"- {error_msg}\n")
            else:
                f.write("- No significant global pipeline errors recorded.\n")
            f.write("\n")

            f.write("## Input Row Attrition Summary:\n")
            if attrition_data_list_for_metrics:
                total_input_rows = metrics.get("data_processing_stats", {}).get("input_rows_count", 0)
                rows_not_yielding_contact = len(attrition_data_list_for_metrics)
                rows_yielding_contact = total_input_rows - rows_not_yielding_contact

                f.write(f"- **Total Input Rows Processed:** {total_input_rows}\n")
                f.write(f"- **Input Rows Yielding at Least One Contact:** {rows_yielding_contact}\n")
                f.write(f"- **Input Rows Not Yielding Any Contact:** {rows_not_yielding_contact}\n\n")

                if rows_not_yielding_contact > 0:
                    f.write("### Reasons for Non-Extraction (Fault Categories):\n")
                    fault_category_counts: Dict[str, int] = {}
                    for item in attrition_data_list_for_metrics:
                        fault = item.get("Determined_Fault_Category", "Unknown")
                        fault_category_counts[fault] = fault_category_counts.get(fault, 0) + 1
                    
                    for fault, count in sorted(fault_category_counts.items()):
                        f.write(f"  - **{fault}:** {count}\n")
                    f.write("\n")

                    # Optional: Top N specific reasons
                    # f.write("### Top Specific Reasons for Non-Extraction:\n")
                    # specific_reason_counts: Dict[str, int] = {}
                    # for item in attrition_data_list_for_metrics:
                    #     reason = item.get("Final_Row_Outcome_Reason", "Unknown")
                    #     specific_reason_counts[reason] = specific_reason_counts.get(reason, 0) + 1
                    # sorted_reasons = sorted(specific_reason_counts.items(), key=lambda x: x[1], reverse=True)
                    # for reason, count in sorted_reasons[:5]: # Top 5
                    #    f.write(f"  - *{reason.replace('_', ' ').title()}:* {count}\n")
                    # f.write("\n")
            else:
                f.write("- No input rows recorded in the attrition report (all rows presumably yielded contacts or failed critically before attrition tracking).\n")
            f.write("\n")

        logger.info(f"Run metrics successfully written to {metrics_file_path}")
    except IOError as e:
        logger.error(f"Failed to write run metrics to {metrics_file_path}: {e}", exc_info=True)
    except Exception as e_global:
        logger.error(f"An unexpected error occurred while writing metrics to {metrics_file_path}: {e_global}", exc_info=True)
def write_canonical_domain_summary_report(
    run_id: str,
    domain_journey_data: Dict[str, Dict[str, Any]],
    output_dir: str
) -> int:
    """
    Writes the canonical domain journey data to an Excel file.
    """
    if not domain_journey_data:
        logger.info("No data for canonical domain summary report. Skipping file creation.")
        return 0

    report_filename = f"canonical_domain_processing_summary_{run_id}.xlsx"
    report_path = os.path.join(output_dir, report_filename)

    report_data_list = []
    for domain, data in domain_journey_data.items():
        row = {"Canonical_Domain": domain}
        row.update(data)
        report_data_list.append(row)
    
    report_df = pd.DataFrame(report_data_list)

    columns_order = [
        "Canonical_Domain", "Input_Row_IDs", "Input_CompanyNames", "Input_GivenURLs",
        "Pathful_URLs_Attempted_List", "Overall_Scraper_Status_For_Domain",
        "Total_Pages_Scraped_For_Domain", "Scraped_Pages_Details_Aggregated",
        "Regex_Candidates_Found_For_Any_Pathful", "LLM_Calls_Made_For_Domain",
        "LLM_Total_Raw_Numbers_Extracted", "LLM_Total_Consolidated_Numbers_Found",
        "LLM_Consolidated_Number_Types_Summary", "LLM_Processing_Error_Encountered_For_Domain",
        "LLM_Error_Messages_Aggregated", "Final_Domain_Outcome_Reason",
        "Primary_Fault_Category_For_Domain"
    ]

    for col in columns_order:
        if col not in report_df.columns:
            report_df[col] = None
            logger.warning(f"Column '{col}' was not found in canonical_domain_summary_report DataFrame and was initialized to None.")
    report_df = report_df[columns_order]

    for col_name in ["Input_Row_IDs", "Input_CompanyNames", "Input_GivenURLs", "Pathful_URLs_Attempted_List", "LLM_Error_Messages_Aggregated"]:
        if col_name in report_df.columns:
            report_df[col_name] = report_df[col_name].apply(lambda x: ", ".join(sorted(list(map(str, x)))) if isinstance(x, (set, list)) else x)
    
    for col_with_counter in ["Scraped_Pages_Details_Aggregated", "LLM_Consolidated_Number_Types_Summary"]:
        if col_with_counter in report_df.columns:
            report_df[col_with_counter] = report_df[col_with_counter].apply(
                lambda x: json.dumps(dict(x)) if isinstance(x, Counter) else x
            )

    try:
        with pd.ExcelWriter(report_path, engine='openpyxl') as writer:
            report_df.to_excel(writer, index=False, sheet_name='Canonical_Domain_Summary')
            worksheet = writer.sheets['Canonical_Domain_Summary']
            for col_idx, col_name in enumerate(report_df.columns):
                series_data = report_df[col_name]
                max_val_len = series_data.astype(str).map(len).max() if not series_data.empty else 0
                column_header_len = len(str(col_name))
                adjusted_width = max(max_val_len, column_header_len) + 5
                worksheet.column_dimensions[get_column_letter(col_idx + 1)].width = adjusted_width
        
        logger.info(f"Canonical domain summary report successfully saved to {report_path}")
        return len(report_df)
    except Exception as e:
        logger.error(f"Failed to write canonical domain summary report to {report_path}: {e}", exc_info=True)
        return 0


if __name__ == '__main__':
    if not logger.hasHandlers(): 
        logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
    
    main()