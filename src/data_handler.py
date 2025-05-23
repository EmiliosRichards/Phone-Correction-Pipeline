import pandas as pd
import phonenumbers
import logging
import uuid # For RunID
from typing import Optional, List, Dict, Any, Union, Iterable # Added Iterable
from urllib.parse import urlparse, urlunparse # Added for URL parsing
import os # Added for os.path.exists and os.makedirs
import re # For new helper functions
from openpyxl.utils import get_column_letter # For auto-sizing columns
from openpyxl import load_workbook # Added for smart Excel reading
import csv # Added for smart CSV reading

# Import AppConfig directly. Its __init__ handles .env loading.
# If this import fails, it's a critical setup error for the application.
from .core.config import AppConfig
from .core.schemas import ( # Added imports for new schemas
    PhoneNumberLLMOutput,
    ConsolidatedPhoneNumberSource,
    ConsolidatedPhoneNumber,
    CompanyContactDetails
)

# Configure logging
# setup_logging() might rely on environment variables loaded by AppConfig's instantiation.
try:
    from .core.logging_config import setup_logging
    # AppConfig() is instantiated globally in config.py if needed by other modules,
    # or when an instance is created. Here, we just ensure logging is set up.
    setup_logging()
    logger = logging.getLogger(__name__)
except ImportError:
    logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    logger = logging.getLogger(__name__)
    logger.info("Basic logging configured due to missing core.logging_config or its dependencies.")


def get_canonical_base_url(url_string: str, log_level_for_non_domain_input: int = logging.WARNING) -> str | None:
    """
    Extracts the canonical base URL (scheme + netloc, with 'www.' removed from netloc)
    from a URL string.
    e.g., "http://www.example.com/path?query" -> "http://example.com"
    e.g., "example.com/path" -> "http://example.com"
    """
    if not url_string or not isinstance(url_string, str):
        logger.warning("get_canonical_base_url received empty or non-string input.")
        return None
    try:
        # Ensure a scheme is present for urlparse to work correctly
        # and handle cases where URL might be missing it (e.g. "www.example.com")
        temp_url = url_string
        if not temp_url.startswith(('http://', 'https://')):
            # Check if it looks like a domain that might have had a scheme stripped
            # or if it's just a path fragment. A simple check for a dot.
            if '.' not in temp_url.split('/')[0]: # if no dot in first part before a slash, it's likely not a domain
                 logger.log(log_level_for_non_domain_input, f"Input string '{url_string}' (when attempting to derive a base URL) does not appear to be a valid absolute URL or domain. This may be an original input value.")
                 return None # Or consider returning the original string if it's a relative path and that's desired
            temp_url = 'http://' + temp_url # Default to http if no scheme

        parsed = urlparse(temp_url)

        if not parsed.netloc:
            logger.log(log_level_for_non_domain_input, f"Could not determine network location (netloc) for input string '{url_string}' (parsed as '{temp_url}' when attempting to derive a base URL).")
            return None

        netloc = parsed.netloc
        # Normalize by removing 'www.' prefix if it exists
        if netloc.startswith('www.'):
            netloc = netloc[4:]

        # Use original scheme if present from parsing temp_url (which had a scheme added if missing)
        # or default to 'http' if somehow scheme is still empty (should not happen with current logic)
        scheme = parsed.scheme if parsed.scheme else 'http'

        # Reconstruct the base URL using urlunparse for proper formatting
        base_url = urlunparse((scheme, netloc, '', '', '', ''))
        return base_url
    except Exception as e:
        logger.error(f"Error parsing URL '{url_string}' to get base URL: {e}", exc_info=True)
        return None

def _is_row_empty(row_values: Iterable) -> bool:
    """Checks if all values in a row are None, empty string, or whitespace-only."""
    if not row_values: # Handles case where row_values itself might be None or an empty list/tuple
        return True
    return all(pd.isna(value) or (isinstance(value, str) and not value.strip()) for value in row_values)

def load_and_preprocess_data(file_path: str, app_config_instance: Optional[AppConfig] = None) -> pd.DataFrame | None:
    """
    Loads data from a CSV or Excel file, standardizes column names, initializes
    new columns required for the pipeline, and applies initial normalization
    to any existing phone numbers. Implements smart reading for open-ended ranges
    to stop after a configured number of consecutive empty rows.
    """
    current_config_instance: AppConfig
    if app_config_instance:
        current_config_instance = app_config_instance
    else:
        current_config_instance = AppConfig()

    skip_rows_val: Optional[int] = None
    nrows_val: Optional[int] = None
    consecutive_empty_rows_to_stop: int = 3 # Default, will be overridden by config

    if hasattr(current_config_instance, 'skip_rows_config'):
        skip_rows_val = current_config_instance.skip_rows_config
    if hasattr(current_config_instance, 'nrows_config'):
        nrows_val = current_config_instance.nrows_config
    if hasattr(current_config_instance, 'consecutive_empty_rows_to_stop'):
        consecutive_empty_rows_to_stop = current_config_instance.consecutive_empty_rows_to_stop

    log_message_parts = []
    if skip_rows_val is not None:
        log_message_parts.append(f"skipping {skip_rows_val} data rows (0-indexed)")
    if nrows_val is not None:
        log_message_parts.append(f"reading {nrows_val} data rows")
    
    smart_read_active = (nrows_val is None and consecutive_empty_rows_to_stop > 0)
    if smart_read_active:
        log_message_parts.append(f"smart read active (stop after {consecutive_empty_rows_to_stop} empty rows)")

    if log_message_parts:
        logger.info(f"Data loading configuration: {', '.join(log_message_parts)}.")
    else:
        logger.info("No specific row range configured; loading all rows (or smart read if enabled).")

    # pandas_skiprows_arg is for when not using smart read or for initial skip in some smart read scenarios
    pandas_skiprows_arg: Union[int, List[int]]
    if skip_rows_val is not None and skip_rows_val > 0:
        pandas_skiprows_arg = list(range(1, skip_rows_val + 1)) # Skips file lines 1 to skip_rows_val
    else:
        pandas_skiprows_arg = 0 # Skip no file lines after header

    df: Optional[pd.DataFrame] = None
    
    try:
        logger.info(f"Attempting to load data from: {file_path}")

        if smart_read_active:
            logger.info(f"Smart read enabled. Max consecutive empty rows: {consecutive_empty_rows_to_stop}")
            header: Optional[List[str]] = None
            data_rows: List[List[Any]] = []
            empty_row_counter = 0
            
            # Effective skip for data rows (0-indexed)
            actual_data_rows_to_skip = skip_rows_val if skip_rows_val is not None else 0

            if file_path.endswith(('.xls', '.xlsx')):
                workbook = load_workbook(filename=file_path, read_only=True, data_only=True)
                sheet = workbook.active
                
                if sheet is None:
                    logger.warning(f"Excel file {file_path} does not have an active sheet or is empty. Returning empty DataFrame.")
                    # Ensure header is None and data_rows is empty so an empty DataFrame is created later
                    header = None
                    data_rows = []
                    # Skip the rest of the Excel processing block
                else:
                    rows_iter = sheet.iter_rows()
                    
                    # 1. Read header
                    try:
                        header_row_values = next(rows_iter)
                        header = [str(cell.value) if cell.value is not None else '' for cell in header_row_values]
                        logger.info(f"Excel header read: {header}")
                    except StopIteration:
                        logger.warning(f"Excel file {file_path} seems empty (no header row after check).")
                        # This path should ideally not be hit if sheet is not None but has no rows.
                        # If it is, header will remain None, data_rows empty.
                        pass # Allow to proceed to df creation with empty data

                    # Only proceed if header was successfully read
                    if header is not None:
                        # 2. Skip initial data rows
                        for _ in range(actual_data_rows_to_skip):
                            try:
                                next(rows_iter)
                            except StopIteration:
                                logger.info(f"Reached end of Excel file while skipping initial {actual_data_rows_to_skip} data rows.")
                                break
                        
                        # 3. Read data with empty row detection
                        for row_idx, row_values_tuple in enumerate(rows_iter):
                            current_row_values = [cell.value for cell in row_values_tuple]
                            if _is_row_empty(current_row_values):
                                empty_row_counter += 1
                                if empty_row_counter >= consecutive_empty_rows_to_stop:
                                    logger.info(f"Stopping Excel read: Found {empty_row_counter} consecutive empty rows at data row index {actual_data_rows_to_skip + row_idx}.")
                                    break
                            else:
                                empty_row_counter = 0
                                data_rows.append(current_row_values)
                
                # This block is now outside the 'else' for sheet is None,
                # but relies on 'header' and 'data_rows' being correctly set.
                if header: # If header is None (e.g. sheet was None or empty), this won't run
                    df = pd.DataFrame(data_rows, columns=header)
                    logger.info(f"Smart read from Excel resulted in {len(data_rows)} data rows.")
                elif not data_rows and header is None: # Case: sheet was None or truly empty
                    df = pd.DataFrame() # Create an empty DataFrame
                    logger.info("Smart read from Excel: sheet was None or empty, created empty DataFrame.")
                # else: # Should not happen if header is None but data_rows has content
                #    df = pd.DataFrame(data_rows) # Fallback if header is somehow None but data exists

            elif file_path.endswith('.csv'):
                with open(file_path, mode='r', encoding='utf-8', newline='') as csvfile: # Specify encoding
                    reader = csv.reader(csvfile)
                    
                    # 1. Read header
                    try:
                        header = next(reader)
                        logger.info(f"CSV header read: {header}")
                    except StopIteration:
                        logger.warning(f"CSV file {file_path} seems empty (no header row).")
                        return pd.DataFrame()

                    # 2. Skip initial data rows
                    for _ in range(actual_data_rows_to_skip):
                        try:
                            next(reader)
                        except StopIteration:
                            logger.info(f"Reached end of CSV file while skipping initial {actual_data_rows_to_skip} data rows.")
                            break
                    
                    # 3. Read data with empty row detection
                    for row_idx, current_row_values in enumerate(reader):
                        if not current_row_values: # Handle completely blank lines from csv.reader
                            is_empty = True
                        else:
                            is_empty = _is_row_empty(current_row_values)

                        if is_empty:
                            empty_row_counter += 1
                            if empty_row_counter >= consecutive_empty_rows_to_stop:
                                logger.info(f"Stopping CSV read: Found {empty_row_counter} consecutive empty rows at data row index {actual_data_rows_to_skip + row_idx}.")
                                break
                        else:
                            empty_row_counter = 0
                            data_rows.append(current_row_values)
                
                if header:
                    df = pd.DataFrame(data_rows, columns=header)
                else: # Should not happen
                    df = pd.DataFrame(data_rows)
                logger.info(f"Smart read from CSV resulted in {len(data_rows)} data rows.")
            else:
                logger.error(f"Unsupported file type for smart read: {file_path}. Please use CSV or Excel.")
                return None
        else: # Original logic (fixed range or smart read disabled)
            logger.info(f"Using standard pandas read. Pandas skiprows argument: {pandas_skiprows_arg}, nrows: {nrows_val}")
            if file_path.endswith('.csv'):
                df = pd.read_csv(file_path, header=0, skiprows=pandas_skiprows_arg, nrows=nrows_val)
            elif file_path.endswith(('.xls', '.xlsx')):
                df = pd.read_excel(file_path, header=0, skiprows=pandas_skiprows_arg, nrows=nrows_val)
            else:
                logger.error(f"Unsupported file type: {file_path}. Please use CSV or Excel.")
                return None
        
        if df is None: # Should only happen if smart read was attempted for unsupported file type
            logger.error(f"DataFrame is None after loading attempt for {file_path}. This indicates an issue with the loading logic.")
            return None

        logger.info(f"Columns loaded: {df.columns.tolist() if df is not None and not df.empty else 'N/A (DataFrame is None or empty)'}")
        
        if df.empty:
            logger.warning(f"Loaded DataFrame from {file_path} is empty. This could be due to an empty input file, all rows being skipped, or smart read stopping early.")
            # If df is empty, we still want to ensure essential columns are present for later stages if they expect them.
            # The new_columns loop later will add them.

        # --- Post-loading processing (rename_map, new_columns, etc.) ---
        rename_map = {
            "Unternehmen": "CompanyName",
            "Webseite": "GivenURL",
            "Telefonnummer": "GivenPhoneNumber",
            "Beschreibung": "Description" # Keep description
        }
        
        actual_rename_map = {k: v for k, v in rename_map.items() if k in df.columns}
        if actual_rename_map:
             df.rename(columns=actual_rename_map, inplace=True)
        logger.info(f"DataFrame columns after renaming: {df.columns.tolist()}")

        new_columns = [
            "NormalizedGivenPhoneNumber", "ScrapingStatus",
            "Overall_VerificationStatus", "Original_Number_Status",
            "Primary_Number_1", "Primary_Type_1", "Primary_SourceURL_1",
            "Secondary_Number_1", "Secondary_Type_1", "Secondary_SourceURL_1",
            "Secondary_Number_2", "Secondary_Type_2", "Secondary_SourceURL_2",
            "RunID", "TargetCountryCodes"
        ]

        current_run_id = str(uuid.uuid4())

        for col in new_columns:
            if col not in df.columns:
                if col == "RunID":
                    df[col] = current_run_id
                elif col == "TargetCountryCodes":
                    # Ensure this is robust for empty df
                    df[col] = pd.Series([["DE", "AT", "CH"] for _ in range(len(df))] if not df.empty else [], dtype=object)
                elif col in ["ScrapingStatus", "Overall_VerificationStatus", "Original_Number_Status"]:
                    df[col] = "Pending"
                elif col.startswith("Primary_") or col.startswith("Secondary_"):
                    df[col] = None 
                else:
                    df[col] = None
        
        logger.info(f"Successfully loaded and structured data from {file_path}. DataFrame shape: {df.shape}")
        
        if "GivenPhoneNumber" in df.columns and not df.empty:
            df = apply_phone_normalization(df.copy(),
                                           phone_column="GivenPhoneNumber",
                                           normalized_column="NormalizedGivenPhoneNumber",
                                           region_column="TargetCountryCodes")
            logger.info("Applied initial phone number normalization.")
        elif "GivenPhoneNumber" not in df.columns:
            logger.warning("'GivenPhoneNumber' column not found. Skipping phone normalization.")
            if "NormalizedGivenPhoneNumber" not in df.columns:
                 df["NormalizedGivenPhoneNumber"] = None
        else: # df is empty
            logger.info("DataFrame is empty, skipping phone normalization.")
            if "NormalizedGivenPhoneNumber" not in df.columns: # Still ensure column exists
                 df["NormalizedGivenPhoneNumber"] = None


        return df
    except FileNotFoundError:
        logger.error(f"Error: The file {file_path} was not found.")
        return None
    except pd.errors.EmptyDataError: # This might be caught by smart read logic earlier for empty files
        logger.error(f"Error: The file {file_path} is empty (pandas EmptyDataError).")
        return pd.DataFrame() # Return empty DataFrame as per original behavior for this specific error
    except Exception as e:
        logger.error(f"An unexpected error occurred while loading data from {file_path}: {e}", exc_info=True)
        return None


def normalize_phone_number(phone_number_str: str, region: str | None = None) -> str | None:
    """
    Normalizes a given phone number string to E.164 format if valid.

    Uses the `python-phonenumbers` library to parse and validate the phone number.
    If a region is provided, it's used as a hint for parsing numbers without
    an international prefix.

    Args:
        phone_number_str (str): The phone number string to normalize.
        region (str | None, optional): A CLDR region code (e.g., "US", "DE")
            to assist in parsing. Defaults to None.

    Returns:
        str | None: The phone number in E.164 format (e.g., "+4930123456")
        if it's valid. Returns "InvalidFormat" if the number cannot be parsed
        or is considered invalid by the library. Returns None if the input
        `phone_number_str` is empty or not a string.

    Raises:
        Logs warnings for parsing errors or invalid numbers.
        Logs errors for unexpected exceptions during normalization.
    """
    if not phone_number_str or not isinstance(phone_number_str, str):
        return None
    try:
        parsed_number = phonenumbers.parse(phone_number_str, region)
        if phonenumbers.is_valid_number(parsed_number):
            return phonenumbers.format_number(parsed_number, phonenumbers.PhoneNumberFormat.E164)
        else:
            logger.info(f"Phone number '{phone_number_str}' (region: {region}) is not valid.")
            return "InvalidFormat"
    except phonenumbers.phonenumberutil.NumberParseException as e:
        logger.info(f"Could not parse phone number '{phone_number_str}' (region: {region}): {e}")
        return "InvalidFormat"
    except Exception as e:
        logger.error(f"Unexpected error normalizing phone number '{phone_number_str}': {e}")
        return "InvalidFormat"


def apply_phone_normalization(df: pd.DataFrame, phone_column: str = "GivenPhoneNumber",
                              normalized_column: str = "NormalizedGivenPhoneNumber",
                              region_column: str | None = None) -> pd.DataFrame:
    """
    Applies phone number normalization to a specified column in a DataFrame.

    Iterates over each row of the DataFrame, takes the phone number from
    `phone_column`, and attempts to normalize it using the
    `normalize_phone_number` function. The result is stored in the
    `normalized_column`. If `region_column` is provided and contains region
    codes (e.g., a list like ['DE', 'CH'] or a single string 'DE'),
    the first region is used as a hint for parsing.

    Args:
        df (pd.DataFrame): The DataFrame to process.
        phone_column (str, optional): The name of the column containing
            the phone numbers to normalize. Defaults to "GivenPhoneNumber".
        normalized_column (str, optional): The name of the column where
            normalized phone numbers will be stored. Defaults to
            "NormalizedGivenPhoneNumber".
        region_column (str | None, optional): The name of the column
            containing region codes (e.g., 'DE', 'US') to aid parsing.
            If the column contains a list of codes, the first one is used.
            Defaults to None.

    Returns:
        pd.DataFrame: The DataFrame with the added `normalized_column`
        containing normalized phone numbers or error strings.

    Raises:
        Logs an error if the `phone_column` is not found in the DataFrame.
    """
    if phone_column not in df.columns:
        logger.error(f"Phone column '{phone_column}' not found in DataFrame.")
        df[normalized_column] = None
        return df

    def normalize_row(row: pd.Series) -> str | None:
        phone_number = row[phone_column]
        default_region: str | None = None
        if region_column and region_column in row and row[region_column]:
            if isinstance(row[region_column], list) and len(row[region_column]) > 0:
                default_region = row[region_column][0]
            elif isinstance(row[region_column], str):
                default_region = row[region_column]
        return normalize_phone_number(str(phone_number), region=default_region)

    df[normalized_column] = df.apply(normalize_row, axis=1) # type: ignore
    logger.info(f"Applied phone normalization to '{phone_column}', results in '{normalized_column}'.")
    return df


def get_classification_priority(classification: str, phone_type: str) -> tuple[int, int]:
    """
    Assigns a numerical priority for sorting.
    Primary key: classification.
    Secondary key: phone_type preference.
    Lower numbers are higher priority.
    """
    classification_priority_map = {
        "Primary": 1,
        "Secondary": 2,
        "Support": 3,
        "Low Relevance": 4,
        "Non-Business": 5,
        "Unknown": 6
    }
    primary_prio = classification_priority_map.get(classification, 99)

    # Define preference for types, especially within the same classification
    # Lower number means it comes earlier in sort (higher preference)
    type_priority_map = {
        # Most preferred types
        "Main Line": 1,
        "Mainline": 1, # Alias
        "Headquarters": 2,
        "Zentrale": 2, # Alias for Headquarters/Main
        "Reception": 3,
        # Departmental / Specific important lines
        "Sales": 10,
        "Sales Department": 10,
        "Customer Service": 11,
        "Support": 12,
        "Support Hotline": 12,
        "Technical Support": 13,
        "Info-Hotline": 15, # Give Info-Hotline a slightly lower preference than direct support/service
        "RA-MICRO Online": 20, # Specific types from example
        "Vertragsmanagement": 21, # Contract Management
        "Direct Dial": 25,
        "Mobile": 30,
        # Less preferred, but still business relevant
        "Fax": 80,
        # Default for unknown/other types
        "Unknown": 99
    }
    # Normalize type for lookup (e.g. lowercase, remove spaces if needed, though current types are fairly clean)
    # For now, direct lookup.
    secondary_prio = type_priority_map.get(phone_type, 90) # Default for types not explicitly listed

    return (primary_prio, secondary_prio)

def process_and_consolidate_contact_data(
    llm_results: List[PhoneNumberLLMOutput],
    company_name_from_input: Optional[str],
    initial_given_url: str
) -> CompanyContactDetails | None:
    """
    Processes a list of LLM-extracted phone numbers for a single company,
    consolidates them by unique number, aggregates their sources (types and paths),
    and groups them under a canonical base URL.

    Args:
        llm_results: A list of PhoneNumberLLMOutput objects, typically all numbers
                     found from scraping pages related to one initial company URL.
        company_name_from_input: The original company name from the input data.
        initial_given_url: The primary URL provided for this company in the input.

    Returns:
        A CompanyContactDetails object if successful, containing the canonical
        base URL, company name, and a list of consolidated phone numbers.
        Returns None if the initial_given_url cannot be processed into a base URL
        or if there are no LLM results to process.
    """
    canonical_base = get_canonical_base_url(initial_given_url)
    if not canonical_base: # This check is crucial. If no base_url, we can't proceed.
        logger.error(f"Could not determine canonical base URL for '{initial_given_url}' (Company: {company_name_from_input}). Cannot consolidate contacts.")
        return None

    logger.info(f"Processing {len(llm_results)} LLM result items for {company_name_from_input or initial_given_url} (Canonical: {canonical_base})")

    if not llm_results:
        logger.info(f"No LLM results provided for {company_name_from_input or initial_given_url} (Canonical: {canonical_base}), returning empty contact details.")
        return CompanyContactDetails(
            company_name=company_name_from_input,
            canonical_base_url=canonical_base, # Use the derived canonical_base
            consolidated_numbers=[],
            original_input_urls=[initial_given_url] if initial_given_url else []
        )

    consolidated_numbers_map: Dict[str, ConsolidatedPhoneNumber] = {}
    all_original_source_urls_for_this_company: set[str] = set()
    if initial_given_url: # Add the initial URL itself
        all_original_source_urls_for_this_company.add(initial_given_url)

    skipped_malformed_count = 0
    for llm_item in llm_results:
        if not llm_item.number or not llm_item.source_url: # Basic check
            logger.warning(f"Skipping LLM item for {company_name_from_input or initial_given_url} (Canonical: {canonical_base}) due to missing number or source_url: {llm_item}")
            skipped_malformed_count += 1
            continue
        
        all_original_source_urls_for_this_company.add(llm_item.source_url)

        # Extract path from the specific source_url of the LLM item
        parsed_source_item_url = urlparse(llm_item.source_url)
        source_path = parsed_source_item_url.path
        if parsed_source_item_url.query:
            source_path += "?" + parsed_source_item_url.query
        if not source_path: # If path is empty (e.g. just domain), use '/'
            source_path = "/"

        current_number_info = ConsolidatedPhoneNumberSource(
            type=llm_item.type,
            source_path=source_path,
            original_full_url=llm_item.source_url,
            original_input_company_name=llm_item.original_input_company_name # Added
        )

        if llm_item.number not in consolidated_numbers_map:
            consolidated_numbers_map[llm_item.number] = ConsolidatedPhoneNumber(
                number=llm_item.number,
                classification=llm_item.classification, # Initial classification
                sources=[current_number_info]
            )
        else:
            # Number already seen, add this new source and update classification if higher priority
            existing_consolidated_number = consolidated_numbers_map[llm_item.number]
            is_duplicate_source = False
            for existing_source in existing_consolidated_number.sources:
                if existing_source.original_full_url == current_number_info.original_full_url and \
                   existing_source.type == current_number_info.type:
                    is_duplicate_source = True
                    break
            if not is_duplicate_source:
                existing_consolidated_number.sources.append(current_number_info)
            
            # Update classification if the new one is "better" (based on tuple comparison)
            current_full_priority = get_classification_priority(llm_item.classification, llm_item.type)
            existing_full_priority = get_classification_priority(existing_consolidated_number.classification, existing_consolidated_number.sources[0].type if existing_consolidated_number.sources else "Unknown") # Use type of first source as representative for existing

            # Python's tuple comparison: (1, 10) < (1, 15) is True; (1, 10) < (2, 1) is True
            if current_full_priority < existing_full_priority:
                existing_consolidated_number.classification = llm_item.classification
                # Note: The 'type' of the ConsolidatedPhoneNumber itself isn't a field.
                # The overall classification is updated. The individual sources retain their original types.
    
    final_consolidated_list = sorted(
        list(consolidated_numbers_map.values()),
        # Sort by classification (primary key) then by the type of the *first source* as a secondary key.
        # This assumes the first source's type is representative enough if multiple sources exist for a number.
        # A more robust way might be to determine a "primary type" for the ConsolidatedPhoneNumber if types differ across sources.
        # For now, using the type from the source that set the best classification, or just the first source.
        # The classification itself is already the "best" one found.
        key=lambda cons_phone: get_classification_priority(cons_phone.classification, cons_phone.sources[0].type if cons_phone.sources else "Unknown")
    )

    if skipped_malformed_count > 0:
        logger.warning(f"{skipped_malformed_count} LLM items were skipped due to being malformed for {company_name_from_input or initial_given_url} (Canonical: {canonical_base}).")

    logger.info(f"Consolidated to {len(final_consolidated_list)} unique phone numbers for {company_name_from_input or initial_given_url} (Canonical: {canonical_base})")

    return CompanyContactDetails(
        company_name=company_name_from_input,
        canonical_base_url=canonical_base,
        consolidated_numbers=final_consolidated_list,
        original_input_urls=sorted(list(all_original_source_urls_for_this_company)) # Store all unique URLs processed
    )

# Helper functions adapted from scripts/process_contacts.py
def _extract_base_domain_for_processed_report(company_name_field: Optional[str]) -> Optional[str]:
    """
    Extracts the base domain from a company name field that might contain a URL.
    Example: "https://wolterskluwer.com - AnNoText" -> "wolterskluwer"
    """
    if pd.isna(company_name_field) or not company_name_field:
        return None
    # Find the first part that looks like a URL
    url_match = re.search(r'https?://[^\s]+', str(company_name_field))
    if url_match:
        url_str = url_match.group(0)
        try:
            parsed_url = urlparse(url_str)
            domain = parsed_url.netloc
            # Remove www. if present
            if domain.startswith('www.'):
                domain = domain[4:]
            domain_parts = domain.split('.')
            if len(domain_parts) > 1:
                # Handles 'example.com' -> 'example', 'example.co.uk' -> 'example'
                if domain_parts[-1] in ['com', 'de', 'org', 'net', 'uk', 'io', 'co', 'eu', 'info', 'biz', 'at', 'ch']: # Expanded TLD list
                    if len(domain_parts) > 2 and domain_parts[-2] in ['co']: # for .co.uk etc.
                         return domain_parts[-3] if len(domain_parts) > 2 else domain_parts[0]
                    return domain_parts[-2] if len(domain_parts) > 1 else domain_parts[0]
                return domain_parts[0] # Fallback if TLD not in simple list
            return domain # If only one part (e.g. localhost)
        except Exception as e:
            logger.warning(f"Error parsing URL '{url_str}' in _extract_base_domain_for_processed_report: {e}")
            return None
    # If no URL found, but the field might be a domain itself (e.g. "example.com")
    # This part is an addition to the original script's logic to handle cases where
    # the company_name_field is just "example.com" without "http://"
    try:
        # Attempt to parse as if it's a netloc
        # Add a scheme to help urlparse identify it as a domain
        temp_url_for_parse = company_name_field
        if not temp_url_for_parse.startswith(('http://', 'https://')):
            # Check if it contains a common TLD, indicating it might be a domain
            if not any(tld in temp_url_for_parse for tld in ['.com', '.de', '.org', '.net', '.io', '.co', '.eu', '.info', '.biz', '.at', '.ch']):
                 # If no common TLD, and no scheme, it's unlikely a domain we can process here.
                 # Return the original field or part of it if it's very long.
                 # For this report, if it's not a clear domain, maybe return None or the original.
                 # The script's original behavior was to return None if no http(s) found.
                 return None # Sticking closer to original script's behavior if no http(s)

            temp_url_for_parse = 'http://' + temp_url_for_parse

        parsed_direct = urlparse(temp_url_for_parse)
        if parsed_direct.netloc:
            domain = parsed_direct.netloc
            if domain.startswith('www.'):
                domain = domain[4:]
            domain_parts = domain.split('.')
            if len(domain_parts) > 1:
                if domain_parts[-1] in ['com', 'de', 'org', 'net', 'uk', 'io', 'co', 'eu', 'info', 'biz', 'at', 'ch']:
                    if len(domain_parts) > 2 and domain_parts[-2] in ['co']:
                         return domain_parts[-3] if len(domain_parts) > 2 else domain_parts[0]
                    return domain_parts[-2] if len(domain_parts) > 1 else domain_parts[0]
                return domain_parts[0]
            return domain
    except Exception: # Catch any error from this alternative parsing
        pass # Fall through to return None
    return None


def _extract_phone_number_for_processed_report(phone_field: Optional[str]) -> Optional[str]:
    """
    Extracts the numerical phone number from a string.
    Example: "+4922332055000 (Main Line) [AnNoText]" -> "+4922332055000"
    Also handles: "0123456789 (Type) [Company]" -> "0123456789"
    """
    if pd.isna(phone_field) or not phone_field:
        return None
    # Regex to capture the part that is likely the number, before any parenthesis.
    # It allows digits, +, spaces, hyphens, and parentheses initially.
    match = re.search(r'([+\d\s()-]+)', str(phone_field))
    if match:
        # Take the matched group, then split by '(', take the first part, strip whitespace.
        # Then remove all non-digit and non-+ characters.
        potential_number_part = match.group(1).split('(')[0].strip()
        phone_num = re.sub(r'[^\d+]', '', potential_number_part)
        return phone_num if phone_num else None
    return None

def _extract_number_type_for_processed_report(phone_field: Optional[str]) -> Optional[str]:
    """
    Extracts the number type from parentheses in a phone string.
    Example: "+4922332055000 (Main Line) [AnNoText]" -> "Main Line"
    """
    if pd.isna(phone_field) or not phone_field:
        return None
    match = re.search(r'\((.*?)\)', str(phone_field)) # Finds content within the first pair of parentheses
    if match:
        return match.group(1).strip()
    return "Unknown" # Default if no type found in parentheses, as per user feedback implicit in script

def generate_processed_contacts_report(
    final_contacts_file_path: str, # Changed from all_company_details
    config: AppConfig,
    run_id: str
) -> None:
    """
    Generates the 'Final Processed Contacts' Excel report based on 'Final Contacts.xlsx'.

    This report contains columns: "Company Name", "URL", "Number", "Number Type",
    and "Number Found At". The data is derived by processing the content of
    the 'Final Contacts.xlsx' file (tertiary report).

    Args:
        final_contacts_file_path (str): Path to the 'Final Contacts.xlsx' file.
        config (AppConfig): The application configuration instance.
        run_id (str): The current run ID for constructing the output path.
    """
    logger.info(f"Starting generation of 'Final Processed Contacts' report from: {final_contacts_file_path}")

    try:
        df_source = pd.read_excel(final_contacts_file_path)
        logger.info(f"Successfully read '{final_contacts_file_path}'. Shape: {df_source.shape}")
    except FileNotFoundError:
        logger.error(f"Error: Source file '{final_contacts_file_path}' not found for 'Final Processed Contacts' report.")
        return
    except Exception as e:
        logger.error(f"Error reading source file '{final_contacts_file_path}': {e}", exc_info=True)
        return

    report_data = []
    for index, row in df_source.iterrows():
        # 'CompanyName' in Final Contacts is like "http://domain.com - Actual Name"
        # The _extract_base_domain_for_processed_report expects this format.
        company_name_input = row.get('CompanyName')
        processed_company_name = _extract_base_domain_for_processed_report(company_name_input)

        # 'CanonicalEntryURL' is the base URL.
        url = row.get('CanonicalEntryURL')

        # 'PhoneNumber_1' in Final Contacts is like "NUMBER (TYPE) [SOURCE_COMPANIES]"
        phone_number_field_1 = row.get('PhoneNumber_1')
        number = _extract_phone_number_for_processed_report(phone_number_field_1)
        number_type = _extract_number_type_for_processed_report(phone_number_field_1)

        # 'SourceURL_1' contains the source URLs for the first number.
        number_found_at = row.get('SourceURL_1')

        # Only add row if a number was successfully extracted, to match script's behavior
        # where rows without numbers (or where extraction fails) might be implicitly skipped.
        # The script `process_contacts.py` appends regardless, but values might be None.
        # For consistency with user expectation of "identical row counts", we should append.
        # If number is None, it will be written as blank in Excel.
        report_data.append({
            "Company Name": processed_company_name,
            "URL": url,
            "Number": number,
            "Number Type": number_type,
            "Number Found At": number_found_at
        })

    if not report_data:
        logger.info("No data processed to write for the 'Final Processed Contacts' report (source df might be empty or all rows resulted in no data).")
        return

    report_df = pd.DataFrame(report_data)
    
    columns_order = ["Company Name", "URL", "Number", "Number Type", "Number Found At"]
    # Ensure all columns exist, add if not, then reorder
    for col in columns_order:
        if col not in report_df.columns:
            report_df[col] = None # Add missing columns with None
    report_df = report_df[columns_order]

    full_path: str = "" # Initialize full_path to ensure it's always bound
    try:
        # Construct the correct output path within the run_id specific directory
        # config.output_base_dir is expected to be an absolute path from AppConfig
        run_specific_output_dir = os.path.join(config.output_base_dir, run_id)
        if not os.path.exists(run_specific_output_dir):
            # This directory should ideally be created by the main pipeline
            logger.warning(f"Run-specific output directory '{run_specific_output_dir}' did not exist. Creating it.")
            os.makedirs(run_specific_output_dir, exist_ok=True)

        # Filename template does not include {run_id} by default
        file_name = config.processed_contacts_report_file_name_template
        # if "{run_id}" in file_name: # This check is good practice but not needed for default template
        #     file_name = file_name.format(run_id=run_id)
        
        full_path = os.path.join(run_specific_output_dir, file_name)
        
        with pd.ExcelWriter(full_path, engine='openpyxl') as writer:
            report_df.to_excel(writer, index=False, sheet_name='Processed_Contacts')
            worksheet = writer.sheets['Processed_Contacts']
            for col_idx, column_name_header in enumerate(report_df.columns):
                series = report_df[column_name_header]
                max_len = 0
                if not series.empty:
                    # Calculate max length of data in the column
                    # Ensure series is treated as string for len calculation
                    # And handle potential NaN values by converting to empty string or specific string
                    # Using .astype(str).map(len) is robust.
                    # For NaNs, astype(str) converts them to 'nan'.
                    # We can replace 'nan' with empty string for length calculation if desired,
                    # or just let it be. The script's example output implies blanks for missing data.
                    
                    # Calculate max length of data values
                    # Convert to string, then get length. Max of these lengths.
                    # Handle if all values are NaN or None after conversion (max of empty sequence error)
                    str_series_lengths = series.astype(str).map(len)
                    if not str_series_lengths.empty:
                         max_len_data = str_series_lengths.max()
                         if pd.notna(max_len_data): # Check if max_len_data itself is not NaN
                            max_len = int(max_len_data)


                # Consider header length
                header_length = len(str(column_name_header))
                adjusted_width = max(max_len, header_length) + 2 # Add padding
                worksheet.column_dimensions[get_column_letter(col_idx + 1)].width = adjusted_width
        
        logger.info(f"'Final Processed Contacts' report generated successfully: {full_path}")
    except Exception as e:
        logger.error(f"Failed to write 'Final Processed Contacts' report to {full_path}: {e}", exc_info=True)

# TODO: [FutureEnhancement] The __main__ block below was for direct script execution, testing, and demonstration.
# It includes logic to create dummy configurations and data if real test files are not found.
# Commented out to prevent execution during normal library use and to avoid creating dummy files.
# It can be uncommented for debugging or understanding standalone script usage.
# if __name__ == '__main__':
#     # Create dummy core.logging_config for direct execution if not present
#     if not os.path.exists("core"):
#         os.makedirs("core")
#     if not os.path.exists("core/logging_config.py"):
#         with open("core/logging_config.py", "w") as f:
#             f.write("import logging\n\ndef setup_logging(level=logging.INFO):\n    logging.basicConfig(level=level, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')\n    # logger = logging.getLogger(__name__) # Avoid getLogger here, let the caller do it.\n    # logger.info('Dummy logging_config setup.')\n") # Removed info log from dummy to avoid confusion
        
#         # If dummy was created, attempt to import and use its setup_logging
#         try:
#             from core.logging_config import setup_logging as main_setup_logging
#             main_setup_logging()
#             logger = logging.getLogger(__name__) # Re-initialize logger with the new setup
#             logger.info("Using dummy logging_config for __main__ block.")
#         except ImportError:
#             logger.error("Failed to import setup_logging from created dummy core.logging_config.")
#             # Fallback to basic config if even dummy import fails, though logger is already basicConfig'd from top
#             pass # logger should already be configured from the top-level except block
    
#     # Path relative to the project root (6 - Phone Correction Full Pipeline)
#     # This script is in phone_validation_pipeline/src/
#     # data_to_be_inputed.xlsx is in the root of the project.
#     # So, from src, it's ../../data_to_be_inputed.xlsx
#     test_file_path = "../../data_to_be_inputed.xlsx"

#     if not os.path.exists(test_file_path):
#         logger.warning(f"Test Excel file not found at {test_file_path}. Creating a dummy CSV for demonstration as fallback.")
#         test_file_path = "../../data_to_be_inputed_dummy.csv" # Save dummy as CSV
#         dummy_data = {
#             'Unternehmen': ['TestFirma AG', 'Beispiel GmbH', 'Muster & Co.'],
#             'Beschreibung': ['Software', 'Beratung', 'Handel'],
#             'Webseite': ['www.test.de', 'http://beispiel.com', 'https://muster.ch'],
#             'Telefonnummer': ['+49 (0) 30 123456', '044 765 43 21', 'invalid-number']
#         }
#         dummy_df = pd.DataFrame(dummy_data)
#         try:
#             dummy_df.to_csv(test_file_path, index=False, encoding='utf-8')
#             logger.info(f"Created dummy CSV at {test_file_path}")
#         except Exception as e:
#             logger.error(f"Could not create dummy CSV for testing: {e}")
#             test_file_path = None

#     if test_file_path:
#         processed_df = load_and_preprocess_data(test_file_path)

#         if processed_df is not None:
#             logger.info("\nProcessed DataFrame head:")
#             logger.info(processed_df.head().to_string())
            
#             logger.info("\nRelevant columns sample:")
#             relevant_cols = ["CompanyName", "GivenURL", "GivenPhoneNumber", "NormalizedGivenPhoneNumber", "TargetCountryCodes", "RunID"]
#             existing_relevant_cols = [col for col in relevant_cols if col in processed_df.columns]
#             logger.info(processed_df[existing_relevant_cols].head(10).to_string())
#         else:
#             logger.error("Failed to load and preprocess data, cannot proceed with example.")
#     else:
#         logger.error("Test file path not set, cannot run example.")